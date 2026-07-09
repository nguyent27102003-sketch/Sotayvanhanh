import os
import re
import sys
import json

sys.stdout.reconfigure(encoding='utf-8')

shopee_md = "src_markdown/shopee_handbook.md"
tiktok_basic_md = "src_markdown/tiktok_basic_handbook.md"
tiktok_advanced_md = "src_markdown/tiktok_advanced_handbook.md"
output_html = "handbook.html"

# Verify files exist
for path in [shopee_md, tiktok_basic_md, tiktok_advanced_md]:
    if not os.path.exists(path):
        print(f"Error: Required markdown file not found at {path}")
        sys.exit(1)

with open(shopee_md, "r", encoding="utf-8") as f:
    shopee_content = f.read()

with open(tiktok_basic_md, "r", encoding="utf-8") as f:
    tiktok_basic_content = f.read()

with open(tiktok_advanced_md, "r", encoding="utf-8") as f:
    tiktok_advanced_content = f.read()

def clean_spaces(text):
    return re.sub(r' {2,}', ' ', text)

shopee_content = clean_spaces(shopee_content)
tiktok_basic_content = clean_spaces(tiktok_basic_content)
tiktok_advanced_content = clean_spaces(tiktok_advanced_content)

# Extract Shopee Products from Excel
def get_shopee_products_list():
    excel_path = "TÍNH CHI PHÍ SHOPPE 2.xlsx"
    if not os.path.exists(excel_path):
        excel_path = r"c:\Users\Administrator\Downloads\TÍNH CHI PHÍ SHOPPE 2.xlsx"
    if not os.path.exists(excel_path):
        print(f"Warning: Shopee Excel not found at {excel_path}, using fallback.")
        return []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb["DATA"]
        products = []
        for r in range(2, sheet.max_row + 1):
            sku = sheet.cell(row=r, column=2).value
            name = sheet.cell(row=r, column=3).value
            brand = sheet.cell(row=r, column=4).value
            category = sheet.cell(row=r, column=5).value
            comm_regular = sheet.cell(row=r, column=8).value
            comm_mall = sheet.cell(row=r, column=9).value
            ref_price = sheet.cell(row=r, column=11).value
            cogs = sheet.cell(row=r, column=12).value
            packing = sheet.cell(row=r, column=13).value
            default_cost = sheet.cell(row=r, column=14).value
            policy = sheet.cell(row=r, column=16).value
            if sku is not None:
                products.append({
                    "sku": str(sku),
                    "name": str(name) if name else "",
                    "brand": str(brand) if brand else "",
                    "category": str(category) if category else "",
                    "comm_regular": float(comm_regular) if comm_regular is not None else 0.1,
                    "comm_mall": float(comm_mall) if comm_mall is not None else 0.12,
                    "ref_price": float(ref_price) if ref_price is not None else 0.0,
                    "cogs": float(cogs) if cogs is not None else None,
                    "packing": float(packing) if packing is not None else 0.0,
                    "default_cost": float(default_cost) if default_cost is not None else 0.0,
                    "policy": str(policy) if policy else ""
                })
        return products
    except Exception as e:
        print(f"Error parsing Shopee Excel: {e}")
        return []

# Extract TikTok Products from Excel
def get_tiktok_products_list():
    excel_path = "TÍNH CHI PHÍ TIKTOK.xlsx"
    if not os.path.exists(excel_path):
        excel_path = r"c:\Users\Administrator\Downloads\TÍNH CHI PHÍ TIKTOK.xlsx"
    if not os.path.exists(excel_path):
        print(f"Warning: TikTok Excel not found at {excel_path}, using fallback.")
        return []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb["DATA PHÍ & SẢN PHẨM"]
        products = []
        for r in range(13, sheet.max_row + 1):
            sku = sheet.cell(row=r, column=1).value
            name = sheet.cell(row=r, column=2).value
            brand = sheet.cell(row=r, column=3).value
            category = sheet.cell(row=r, column=4).value
            comm_regular = sheet.cell(row=r, column=6).value
            comm_mall = sheet.cell(row=r, column=7).value
            cogs = sheet.cell(row=r, column=8).value
            packing = sheet.cell(row=r, column=9).value
            gift = sheet.cell(row=r, column=10).value
            if sku is not None:
                products.append({
                    "sku": str(sku),
                    "name": str(name) if name else "",
                    "brand": str(brand) if brand else "",
                    "category": str(category) if category else "",
                    "comm_regular": float(comm_regular) if comm_regular is not None else 0.1,
                    "comm_mall": float(comm_mall) if comm_mall is not None else 0.12,
                    "cogs": float(cogs) if cogs is not None else 0.0,
                    "packing": float(packing) if packing is not None else 0.0,
                    "gift": float(gift) if gift is not None else 0.0
                })
        return products
    except Exception as e:
        print(f"Error parsing TikTok Excel: {e}")
        return []

shopee_products_db = get_shopee_products_list()
tiktok_products_db = get_tiktok_products_list()

# Fallback datasets if files were missing
if not shopee_products_db:
    shopee_products_db = [
        {"sku": "HC011", "name": "Metacare 1+ 850g", "brand": "Nutricare/MetaCare", "category": "Sữa công thức", "comm_regular": 0.11, "comm_mall": 0.125, "ref_price": 335000, "cogs": None, "packing": 6000, "default_cost": 240500, "policy": "1-2 tuổi"},
        {"sku": "HC012", "name": "Metacare 2+ 850g", "brand": "Nutricare/MetaCare", "category": "Sữa công thức", "comm_regular": 0.11, "comm_mall": 0.125, "ref_price": 345000, "cogs": None, "packing": 6000, "default_cost": 247500, "policy": "2-10 tuổi"},
        {"sku": "HC019", "name": "Hanie Kid pha sẵn - thùng 48 hộp", "brand": "Nutricare/Hanie Kid", "category": "Sữa pha sẵn", "comm_regular": 0.115, "comm_mall": 0.145, "ref_price": 460800, "cogs": None, "packing": 5000, "default_cost": 327560, "policy": "Từ 1 tuổi"},
        {"sku": "HC023", "name": "Nutricare Gold 850g", "brand": "Nutricare", "category": "Sức Khỏe", "comm_regular": 0.17, "comm_mall": 0.19, "ref_price": 521000, "cogs": None, "packing": 6000, "default_cost": 370700, "policy": "Dinh dưỡng y học"}
    ]
if not tiktok_products_db:
    tiktok_products_db = [
        {"sku": "HC001", "name": "Sữa công thức tăng trưởng - mẫu", "brand": "Nutricare/Metacare", "category": "Sữa công thức", "comm_regular": 0.1, "comm_mall": 0.115, "cogs": 300000, "packing": 5000, "gift": 0},
        {"sku": "HC002", "name": "Sữa công thức sơ sinh - mẫu", "brand": "Nutricare/Metacare", "category": "Sữa công thức", "comm_regular": 0.095, "comm_mall": 0.105, "cogs": 320000, "packing": 5000, "gift": 0},
        {"sku": "HC011", "name": "Yến sào - mẫu", "brand": "Nunest", "category": "Yến sào", "comm_regular": 0.125, "comm_mall": 0.157, "cogs": 180000, "packing": 5000, "gift": 0}
    ]

# HTML Head Template
template_head = """<!DOCTYPE html>
<html lang="vi" data-theme="dark" data-platform="shopee">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Học viện Vận hành Thương mại điện tử Việt Nam</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
    <style>
        :root {
            /* General Theme Variables */
            --bg-primary: #0f172a;
            --bg-secondary: #1e293b;
            --bg-sidebar: #0b0f19;
            --text-primary: #f8fafc;
            --text-secondary: #94a3b8;
            --border-color: #334155;
            --card-bg: #1e293b;
            --alert-info-bg: rgba(59, 130, 246, 0.15);
            --alert-info-border: #3b82f6;
            --alert-warning-bg: rgba(245, 158, 11, 0.15);
            --alert-warning-border: #f59e0b;
            --alert-danger-bg: rgba(239, 68, 68, 0.15);
            --alert-danger-border: #ef4444;

            /* Platform Custom Themes */
            --shopee-color: #ee4d2d;
            --shopee-hover: #ff6546;
            --tiktok-color: #25f4ee;
            --tiktok-accent: #fe2c55;
            
            /* Dynamic active variables (switched by platform attribute) */
            --accent-color: var(--shopee-color);
            --accent-hover: var(--shopee-hover);
            --active-item-bg: rgba(238, 77, 45, 0.15);
            --checklist-bg: rgba(238, 77, 45, 0.05);
        }

        /* Adjust colors based on active platform */
        [data-platform="shopee"] {
            --accent-color: var(--shopee-color);
            --accent-hover: var(--shopee-hover);
            --active-item-bg: rgba(238, 77, 45, 0.15);
            --checklist-bg: rgba(238, 77, 45, 0.05);
        }

        [data-platform="tiktok"] {
            --accent-color: var(--tiktok-accent);
            --accent-hover: #ff476f;
            --active-item-bg: rgba(254, 44, 85, 0.15);
            --checklist-bg: rgba(254, 44, 85, 0.05);
        }

        /* Light Mode Adjustments */
        [data-theme="light"] {
            --bg-primary: #f8fafc;
            --bg-secondary: #ffffff;
            --bg-sidebar: #f1f5f9;
            --text-primary: #0f172a;
            --text-secondary: #64748b;
            --border-color: #e2e8f0;
            --card-bg: #ffffff;
            --alert-info-bg: rgba(59, 130, 246, 0.05);
            --alert-info-border: #3b82f6;
            --alert-warning-bg: rgba(245, 158, 11, 0.05);
            --alert-warning-border: #f59e0b;
            --alert-danger-bg: rgba(239, 68, 68, 0.05);
            --alert-danger-border: #ef4444;
        }

        [data-theme="light"][data-platform="shopee"] {
            --active-item-bg: rgba(238, 77, 45, 0.1);
            --checklist-bg: rgba(238, 77, 45, 0.02);
        }

        [data-theme="light"][data-platform="tiktok"] {
            --active-item-bg: rgba(254, 44, 85, 0.1);
            --checklist-bg: rgba(254, 44, 85, 0.02);
        }

        * {
            box-sizing: border-box;
            margin: 0;
            padding: 0;
            font-family: 'Inter', sans-serif;
        }

        body {
            background-color: var(--bg-primary);
            color: var(--text-primary);
            display: flex;
            height: 100vh;
            overflow: hidden;
            transition: background-color 0.3s, color 0.3s;
        }

        /* Sidebar Styling */
        .sidebar {
            width: 320px;
            background-color: var(--bg-sidebar);
            border-right: 1px solid var(--border-color);
            display: flex;
            flex-direction: column;
            height: 100%;
            flex-shrink: 0;
        }

        .sidebar-header {
            padding: 20px;
            border-bottom: 1px solid var(--border-color);
        }

        /* Platform Switcher Tabs */
        .platform-tabs {
            display: flex;
            background: var(--bg-primary);
            padding: 4px;
            border-radius: 8px;
            margin-bottom: 16px;
            border: 1px solid var(--border-color);
        }

        .platform-tab-btn {
            flex: 1;
            padding: 8px 12px;
            background: transparent;
            border: none;
            color: var(--text-secondary);
            font-size: 13px;
            font-weight: 600;
            cursor: pointer;
            border-radius: 6px;
            transition: all 0.2s;
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 6px;
        }

        .platform-tab-btn.active {
            color: #ffffff;
        }

        [data-platform="shopee"] #tab-shopee {
            background: var(--shopee-color);
            color: #ffffff;
        }

        [data-platform="tiktok"] #tab-tiktok {
            background: var(--tiktok-accent);
            color: #ffffff;
        }

        /* Level Tabs for TikTok */
        .level-tabs {
            display: flex;
            background: var(--bg-primary);
            padding: 3px;
            border-radius: 6px;
            border: 1px solid var(--border-color);
        }

        .level-tab-btn {
            flex: 1;
            padding: 6px 8px;
            background: transparent;
            border: none;
            color: var(--text-secondary);
            font-size: 11px;
            font-weight: 600;
            cursor: pointer;
            border-radius: 4px;
            transition: all 0.2s;
            text-align: center;
        }

        .level-tab-btn.active {
            background: var(--border-color);
            color: var(--text-primary);
        }
        
        [data-platform="tiktok"] #btn-tiktok-basic.active {
            background: var(--tiktok-accent);
            color: #ffffff;
        }
        
        /* Low profile styling for advanced button to not make it prominent */
        [data-platform="tiktok"] #btn-tiktok-advanced.active {
            background: var(--bg-secondary);
            border: 1px solid var(--border-color);
            color: var(--text-primary);
        }

        .search-box {
            position: relative;
            margin-bottom: 12px;
        }

        .search-box input {
            width: 100%;
            padding: 10px 14px 10px 36px;
            background-color: var(--bg-primary);
            border: 1px solid var(--border-color);
            border-radius: 8px;
            color: var(--text-primary);
            font-size: 14px;
            outline: none;
            transition: border-color 0.2s;
        }

        .search-box input:focus {
            border-color: var(--accent-color);
        }

        .search-box svg {
            position: absolute;
            left: 12px;
            top: 50%;
            transform: translateY(-50%);
            color: var(--text-secondary);
            width: 16px;
            height: 16px;
        }

        .theme-toggle-btn {
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 8px;
            width: 100%;
            padding: 10px;
            background: var(--bg-secondary);
            border: 1px solid var(--border-color);
            border-radius: 8px;
            color: var(--text-primary);
            font-size: 13px;
            cursor: pointer;
            font-weight: 500;
            transition: background-color 0.2s;
        }

        .theme-toggle-btn:hover {
            background-color: var(--border-color);
        }

        .sidebar-menu {
            flex: 1;
            overflow-y: auto;
            padding: 16px;
        }

        .menu-title {
            font-size: 11px;
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 0.05em;
            color: var(--text-secondary);
            margin: 12px 0 8px 8px;
        }

        .menu-list {
            list-style: none;
        }

        .menu-item {
            margin-bottom: 4px;
        }

        .menu-link {
            display: flex;
            align-items: center;
            padding: 10px 12px;
            color: var(--text-secondary);
            text-decoration: none;
            border-radius: 6px;
            font-size: 13.5px;
            font-weight: 500;
            transition: all 0.2s;
            cursor: pointer;
            line-height: 1.3;
        }

        .menu-link:hover, .menu-link.active {
            color: var(--text-primary);
            background-color: var(--active-item-bg);
        }

        .menu-link.active {
            border-left: 3px solid var(--accent-color);
            border-top-left-radius: 0;
            border-bottom-left-radius: 0;
            font-weight: 600;
        }

        /* Shopee Sidebar Submenu Accordion */
        .menu-header-wrapper {
            display: flex;
            justify-content: space-between;
            align-items: center;
            cursor: pointer;
            border-radius: 6px;
            transition: all 0.2s;
        }
        
        .menu-header-wrapper:hover {
            background-color: var(--active-item-bg);
        }

        .menu-header-wrapper .menu-link {
            flex: 1;
            background: transparent !important;
            padding-right: 4px;
            border-left: none !important;
        }
        
        .menu-header-wrapper:has(.menu-link.active) {
            background-color: var(--active-item-bg);
            border-left: 3px solid var(--accent-color);
            border-top-left-radius: 0;
            border-bottom-left-radius: 0;
        }
        
        .menu-header-wrapper:has(.menu-link.active) .menu-link {
            font-weight: 600;
            color: var(--text-primary);
        }

        .submenu-arrow {
            padding: 8px 12px;
            color: var(--text-secondary);
            font-size: 10px;
            transition: transform 0.2s ease;
            user-select: none;
        }

        .submenu-list {
            list-style: none;
            margin: 2px 0 6px 16px;
            padding-left: 0;
            border-left: 1px solid var(--border-color);
        }

        .submenu-item {
            margin-bottom: 2px;
        }

        .submenu-link {
            display: block;
            padding: 6px 12px;
            color: var(--text-secondary);
            text-decoration: none;
            font-size: 12.5px;
            border-radius: 4px;
            cursor: pointer;
            transition: all 0.2s;
            line-height: 1.3;
        }

        .submenu-link:hover, .submenu-link.active {
            color: var(--text-primary);
            background-color: var(--active-item-bg);
        }
        
        .submenu-link.active {
            font-weight: 600;
            color: var(--shopee-color) !important;
        }

        /* Content Area */
        .content-container {
            flex: 1;
            display: flex;
            flex-direction: column;
            height: 100%;
            background-color: var(--bg-primary);
            overflow-y: auto;
        }

        .content-header {
            padding: 16px 40px;
            border-bottom: 1px solid var(--border-color);
            display: flex;
            justify-content: space-between;
            align-items: center;
            background-color: var(--bg-secondary);
            position: sticky;
            top: 0;
            z-index: 10;
        }

        .content-body {
            padding: 40px;
            max-width: 960px;
            margin: 0 auto;
            width: 100%;
        }

        /* Typography */
        h1, h2, h3, h4 {
            color: var(--text-primary);
            margin-top: 32px;
            margin-bottom: 16px;
            font-weight: 700;
            line-height: 1.3;
        }

        h1.handbook-section-h1 {
            font-size: 26px;
            border-bottom: 2px solid var(--border-color);
            padding-bottom: 12px;
            color: var(--accent-color);
            margin-top: 48px;
        }

        h2.handbook-section-h2 {
            font-size: 20px;
            border-bottom: 1px solid var(--border-color);
            padding-bottom: 8px;
            margin-top: 36px;
        }

        h3.handbook-section-h3 {
            font-size: 17px;
            margin-top: 28px;
        }

        h4.handbook-section-h4 {
            font-size: 15px;
            margin-top: 20px;
            color: var(--text-secondary);
        }

        p, li {
            font-size: 14.5px;
            line-height: 1.6;
            color: var(--text-primary);
            margin-bottom: 12px;
        }

        ul, ol {
            margin-bottom: 20px;
            padding-left: 24px;
        }

        li {
            margin-bottom: 6px;
        }

        strong {
            font-weight: 600;
            color: var(--text-primary);
        }

        code {
            font-family: monospace;
            background: rgba(238, 77, 45, 0.1);
            color: var(--accent-color);
            padding: 2px 6px;
            border-radius: 4px;
            font-size: 13.5px;
        }

        /* Image styling */
        img.handbook-img {
            max-width: 100%;
            height: auto;
            border-radius: 8px;
            border: 1px solid var(--border-color);
            margin: 20px 0;
            cursor: zoom-in;
            transition: transform 0.2s, box-shadow 0.2s;
            display: block;
        }

        img.handbook-img:hover {
            transform: scale(1.01);
            box-shadow: 0 10px 15px -3px rgba(0, 0, 0, 0.3);
        }

        /* Alerts */
        .alert-box {
            padding: 16px 20px;
            border-left: 4px solid transparent;
            border-radius: 6px;
            margin: 20px 0;
        }

        .alert-info {
            background-color: var(--alert-info-bg);
            border-left-color: var(--alert-info-border);
        }

        .alert-warning {
            background-color: var(--alert-warning-bg);
            border-left-color: var(--alert-warning-border);
        }

        .alert-danger {
            background-color: var(--alert-danger-bg);
            border-left-color: var(--alert-danger-border);
        }

        .alert-title {
            font-weight: 700;
            margin-bottom: 8px;
            display: flex;
            align-items: center;
            gap: 8px;
        }

        /* Tables */
        table {
            width: 100%;
            border-collapse: collapse;
            margin: 24px 0;
            background-color: var(--bg-secondary);
            border-radius: 8px;
            overflow: hidden;
            border: 1px solid var(--border-color);
        }

        th, td {
            padding: 12px 16px;
            text-align: left;
            border-bottom: 1px solid var(--border-color);
            font-size: 13.5px;
        }

        th {
            background-color: rgba(238, 77, 45, 0.05);
            color: var(--accent-color);
            font-weight: 600;
        }

        [data-platform="tiktok"] th {
            background-color: rgba(254, 44, 85, 0.05);
        }

        tr:last-child td {
            border-bottom: none;
        }

        tr:hover td {
            background-color: rgba(255, 255, 255, 0.02);
        }

        blockquote {
            border-left: 4px solid var(--border-color);
            padding: 8px 16px;
            margin: 16px 0;
            font-style: italic;
            color: var(--text-secondary);
        }

        /* Widgets */
        .widget-card {
            background-color: var(--bg-secondary);
            border: 1px solid var(--border-color);
            border-radius: 12px;
            padding: 24px;
            margin: 32px 0;
            box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
        }

        .widget-title {
            color: var(--accent-color);
            font-size: 17px;
            font-weight: 600;
            margin-bottom: 16px;
            display: flex;
            align-items: center;
            gap: 8px;
            border-bottom: 1px solid var(--border-color);
            padding-bottom: 8px;
        }

        .calculator-grid {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 20px;
        }

        @media (max-width: 768px) {
            .calculator-grid {
                grid-template-columns: 1fr;
            }
        }

        .form-group {
            margin-bottom: 16px;
        }

        .form-group label {
            display: block;
            font-size: 13.5px;
            font-weight: 500;
            color: var(--text-secondary);
            margin-bottom: 8px;
        }

        .form-group input, .form-group select {
            width: 100%;
            padding: 10px 14px;
            background-color: var(--bg-primary);
            border: 1px solid var(--border-color);
            border-radius: 8px;
            color: var(--text-primary);
            font-size: 14.5px;
            outline: none;
            transition: border-color 0.2s;
        }

        .form-group input:focus, .form-group select:focus {
            border-color: var(--accent-color);
        }

        .calc-result-box {
            background-color: rgba(238, 77, 45, 0.02);
            border: 1px dashed var(--border-color);
            border-radius: 8px;
            padding: 20px;
            display: flex;
            flex-direction: column;
            justify-content: center;
        }

        .result-row {
            display: flex;
            justify-content: space-between;
            margin-bottom: 10px;
            font-size: 13.5px;
        }

        .result-row.total {
            font-size: 18px;
            font-weight: 700;
            color: var(--accent-color);
            border-top: 1px solid var(--border-color);
            padding-top: 10px;
            margin-top: 10px;
            margin-bottom: 0;
        }

        .math-block {
            background: var(--bg-secondary);
            border: 1px solid var(--border-color);
            border-radius: 8px;
            padding: 16px;
            margin: 16px 0;
            text-align: center;
            font-family: 'Courier New', Courier, monospace;
            font-weight: 600;
            color: var(--accent-color);
            font-size: 15px;
        }

        /* Checklist */
        .checklist-container {
            background-color: var(--checklist-bg);
            border: 1px solid var(--border-color);
            border-radius: 12px;
            padding: 24px;
            margin: 24px 0;
        }

        .checklist-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 16px;
            border-bottom: 1px solid var(--border-color);
            padding-bottom: 12px;
        }

        .checklist-progress-bar-bg {
            background-color: var(--border-color);
            height: 8px;
            width: 100%;
            border-radius: 4px;
            margin-bottom: 20px;
            overflow: hidden;
        }

        .checklist-progress-bar-fill {
            background-color: var(--accent-color);
            height: 100%;
            width: 0%;
            border-radius: 4px;
            transition: width 0.3s ease;
        }

        .checklist-item {
            display: flex;
            align-items: flex-start;
            gap: 12px;
            margin-bottom: 12px;
            cursor: pointer;
        }

        .checklist-item input[type="checkbox"] {
            margin-top: 4px;
            width: 16px;
            height: 16px;
            accent-color: var(--accent-color);
            cursor: pointer;
        }

        .checklist-item-text {
            font-size: 14.5px;
            line-height: 1.4;
        }

        .checklist-item.checked .checklist-item-text {
            text-decoration: line-through;
            color: var(--text-secondary);
        }

        /* Platform Switching Content Blocks */
        .platform-content {
            display: none;
        }

        .platform-content.active {
            display: block;
        }

        /* Policy word checker widget */
        .policy-checker-textarea {
            width: 100%;
            height: 100px;
            background-color: var(--bg-primary);
            border: 1px solid var(--border-color);
            border-radius: 8px;
            color: var(--text-primary);
            padding: 12px;
            font-size: 14px;
            resize: vertical;
            outline: none;
            margin-bottom: 12px;
        }

        .policy-checker-textarea:focus {
            border-color: var(--accent-color);
        }

        .policy-result-card {
            background-color: var(--bg-primary);
            border: 1px solid var(--border-color);
            border-radius: 8px;
            padding: 16px;
            margin-top: 12px;
        }

        .policy-tag {
            display: inline-block;
            padding: 2px 8px;
            border-radius: 4px;
            font-size: 12px;
            font-weight: 600;
            margin-right: 6px;
            margin-bottom: 6px;
        }

        .policy-tag.danger {
            background-color: rgba(239, 68, 68, 0.2);
            color: #ef4444;
            border: 1px solid #ef4444;
        }

        .policy-tag.safe {
            background-color: rgba(34, 197, 94, 0.2);
            color: #22c55e;
            border: 1px solid #22c55e;
        }

        /* Lightbox modal */
        .lightbox-modal {
            display: none;
            position: fixed;
            z-index: 100;
            left: 0;
            top: 0;
            width: 100%;
            height: 100%;
            background-color: rgba(15, 23, 42, 0.95);
            align-items: center;
            justify-content: center;
            flex-direction: column;
            opacity: 0;
            transition: opacity 0.3s ease;
        }

        .lightbox-modal.show {
            display: flex;
            opacity: 1;
        }

        .lightbox-content {
            max-width: 90%;
            max-height: 80%;
            border-radius: 8px;
            box-shadow: 0 25px 50px -12px rgba(0, 0, 0, 0.5);
            border: 2px solid var(--border-color);
        }

        .lightbox-caption {
            color: #f8fafc;
            margin-top: 16px;
            font-size: 16px;
            font-weight: 500;
            text-align: center;
        }

        .lightbox-close {
            position: absolute;
            top: 24px;
            right: 32px;
            color: #94a3b8;
            font-size: 40px;
            font-weight: bold;
            cursor: pointer;
            transition: color 0.2s;
        }

        .lightbox-close:hover {
            color: #f8fafc;
        }

        /* Configurator specific styles */
        .profile-btn-group {
            display: flex;
            gap: 10px;
            margin-bottom: 16px;
        }

        .profile-btn {
            flex: 1;
            padding: 10px;
            background-color: var(--bg-primary);
            border: 1px solid var(--border-color);
            border-radius: 8px;
            color: var(--text-secondary);
            font-size: 12.5px;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.2s;
        }

        .profile-btn:hover, .profile-btn.active {
            border-color: var(--accent-color);
            color: var(--text-primary);
            background-color: rgba(238, 77, 45, 0.1);
        }

        [data-platform="tiktok"] .profile-btn:hover, [data-platform="tiktok"] .profile-btn.active {
            background-color: rgba(254, 44, 85, 0.1);
        }

        .obs-config-table {
            width: 100%;
            margin: 12px 0 0 0;
            border-collapse: collapse;
        }

        .obs-config-table td {
            padding: 8px 12px;
            border-bottom: none;
        }

        .copy-btn {
            margin-top: 12px;
            padding: 8px 12px;
            background-color: var(--accent-color);
            border: none;
            border-radius: 6px;
            color: white;
            font-weight: 600;
            cursor: pointer;
            font-size: 13px;
            transition: background-color 0.2s;
            display: inline-block;
            text-align: center;
        }

        .copy-btn:hover {
            opacity: 0.9;
        }

        .dispute-result {
            background-color: var(--bg-primary);
            border: 1px solid var(--border-color);
            border-radius: 8px;
            padding: 16px;
            margin-top: 16px;
        }

        .dispute-result-title {
            font-weight: 600;
            color: var(--accent-color);
            margin-bottom: 8px;
            font-size: 13.5px;
        }

        .template-textarea {
            width: 100%;
            height: 120px;
            background-color: var(--bg-secondary);
            border: 1px solid var(--border-color);
            border-radius: 6px;
            color: var(--text-primary);
            padding: 10px;
            font-size: 13px;
            font-family: inherit;
            resize: none;
            margin-top: 8px;
            outline: none;
        }

        .template-textarea:focus {
            border-color: var(--accent-color);
        }

        /* Scrollbar */
        ::-webkit-scrollbar {
            width: 8px;
            height: 8px;
        }

        ::-webkit-scrollbar-track {
            background: transparent;
        }

        ::-webkit-scrollbar-thumb {
            background: var(--border-color);
            border-radius: 4px;
        }

        ::-webkit-scrollbar-thumb:hover {
            background: var(--text-secondary);
        }

        /* Modal Overlays */
        .modal-overlay {
            display: none;
            position: fixed;
            z-index: 200;
            left: 0;
            top: 0;
            width: 100%;
            height: 100%;
            background-color: rgba(15, 23, 42, 0.7);
            align-items: center;
            justify-content: center;
        }
        
        .modal-overlay.show {
            display: flex;
        }
        
        .modal-card {
            background-color: var(--bg-secondary);
            border: 1px solid var(--border-color);
            border-radius: 12px;
            padding: 24px;
            width: 90%;
            max-width: 450px;
            box-shadow: 0 20px 25px -5px rgba(0, 0, 0, 0.3);
        }
        
        .modal-header {
            font-size: 16px;
            font-weight: 700;
            margin-bottom: 16px;
            color: var(--accent-color);
            border-bottom: 1px solid var(--border-color);
            padding-bottom: 8px;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }

        .modal-close-btn {
            background: none;
            border: none;
            color: var(--text-secondary);
            font-size: 20px;
            font-weight: bold;
            cursor: pointer;
        }

        .modal-footer {
            display: flex;
            justify-content: flex-end;
            gap: 10px;
            margin-top: 20px;
        }

        .btn-secondary {
            background: transparent;
            border: 1px solid var(--border-color);
            color: var(--text-primary);
            padding: 8px 16px;
            border-radius: 6px;
            cursor: pointer;
            font-weight: 600;
            font-size: 13px;
        }

        .btn-primary {
            background: var(--accent-color);
            border: none;
            color: white;
            padding: 8px 16px;
            border-radius: 6px;
            cursor: pointer;
            font-weight: 600;
            font-size: 13px;
        }
    </style>
</head>
<body>
"""

# Widgets Templates
obs_configurator_html = """
<div class="widget-card">
    <div class="widget-title">
        ⚙️ BỘ THIẾT LẬP THÔNG SỐ OBS LIVESTREAM ĐỀ XUẤT NHANH
    </div>
    <p style="font-size: 13px; color: var(--text-secondary); margin-bottom: 12px;">Chọn cấu hình máy tính của bạn để hiển thị thông số OBS tương ứng:</p>
    <div class="profile-btn-group">
        <button class="profile-btn active" onclick="selectOBSProfile('low')">Cấu hình Thấp / Văn phòng</button>
        <button class="profile-btn" onclick="selectOBSProfile('mid')">Cấu hình Trung bình / Cơ bản</button>
        <button class="profile-btn" onclick="selectOBSProfile('high')">Cấu hình Cao / Stream chuyên nghiệp</button>
    </div>
    <div style="background-color: var(--bg-primary); border: 1px solid var(--border-color); border-radius: 8px; padding: 12px;">
        <table class="obs-config-table" style="margin:0; border:none; background:none;">
            <tr>
                <td style="width: 40%; font-weight: 600; color: var(--text-secondary); padding: 6px 12px;">Độ phân giải (Output):</td>
                <td id="obs-res" style="font-weight:600; padding: 6px 12px;">540p (960 x 540) hoặc 720p</td>
            </tr>
            <tr>
                <td style="font-weight: 600; color: var(--text-secondary); padding: 6px 12px;">Khung hình (FPS):</td>
                <td id="obs-fps" style="font-weight:600; padding: 6px 12px;">30 FPS</td>
            </tr>
            <tr>
                <td style="font-weight: 600; color: var(--text-secondary); padding: 6px 12px;">Bitrate Video:</td>
                <td id="obs-bitrate" style="font-weight:600; padding: 6px 12px;">1000 - 1500 Kbps</td>
            </tr>
            <tr>
                <td style="font-weight: 600; color: var(--text-secondary); padding: 6px 12px;">Bộ mã hóa (Encoder):</td>
                <td id="obs-encoder" style="font-weight:600; padding: 6px 12px;">x264 (Phần mềm CPU)</td>
            </tr>
            <tr>
                <td style="font-weight: 600; color: var(--text-secondary); padding: 6px 12px;">Keyframe Interval:</td>
                <td style="font-weight:600; color: var(--shopee-color); padding: 6px 12px;">2 giây (Bắt buộc)</td>
            </tr>
        </table>
    </div>
    <button class="copy-btn" onclick="copyOBSSettings()">Sao chép thông số cấu hình này</button>
</div>
"""

dispute_assistant_html = """
<div class="widget-card">
    <div class="widget-title">
        🛡️ TRỢ LÝ GIẢI QUYẾT TRANH CHẤP TRẢ HÀNG / HOÀN TIỀN
    </div>
    <p style="font-size: 13px; color: var(--text-secondary); margin-bottom: 12px;">Chọn lý do người mua gửi yêu cầu hoàn tiền để chuẩn bị bằng chứng và soạn thảo nội dung khiếu nại gửi Shopee:</p>
    
    <div class="form-group">
        <label for="dispute-reason-select">Lý do Trả hàng/Hoàn tiền của khách:</label>
        <select id="dispute-reason-select" onchange="generateDisputeTemplate()">
            <option value="empty_box">1. Người mua nhận được hộp rỗng / Thất lạc hàng bên trong</option>
            <option value="missing">2. Giao thiếu hàng (Thiếu quà tặng / Thiếu số lượng)</option>
            <option value="damaged">3. Sản phẩm bị bể vỡ / Hư hỏng trong quá trình vận chuyển</option>
            <option value="wrong_item">4. Khách trả lại hàng không đúng mẫu mã đã giao (Đổi tráo hàng)</option>
        </select>
    </div>
    
    <div class="dispute-result">
        <div class="dispute-result-title">📁 Chứng cứ bắt buộc phải có để gửi Shopee:</div>
        <ul id="dispute-evidence-list" style="margin-bottom: 12px; padding-left: 20px; font-size:13px;">
            <!-- Filled by JS -->
        </ul>
        
        <div class="dispute-result-title">📄 Nội dung khiếu nại mẫu (Bằng tiếng Việt):</div>
        <textarea class="template-textarea" id="dispute-text-template" readonly onclick="this.select()"></textarea>
        <p style="font-size: 11px; color: var(--text-secondary); margin-top: 4px;">*Nhấp vào ô văn bản trên để chọn tất cả và sao chép. Hãy điền các thông tin trong ngoặc [ ] trước khi gửi.</p>
    </div>
</div>
"""

tiktok_policy_checker_html = """
<div class="widget-card">
    <div class="widget-title">
        🛡️ BỘ QUÉT TỪ KHÓA AN TOÀN CHÍNH SÁCH LIVESTREAM TIKTOK
    </div>
    <p style="font-size: 13px; color: var(--text-secondary); margin-bottom: 12px;">Nhập kịch bản live hoặc câu nói tư vấn của bạn vào ô dưới đây để kiểm tra xem có chứa từ khóa nhạy cảm/bị cấm của TikTok Shop hay không:</p>
    <textarea class="policy-checker-textarea" id="policy-checker-input" placeholder="Ví dụ: Sản phẩm này tốt nhất thị trường, cam kết 100% giúp bé trị dứt điểm táo bón sau 3 ngày..." onkeyup="checkTikTokPolicy()"></textarea>
    
    <div class="policy-result-card" id="policy-checker-result-box" style="display:none;">
        <div style="font-weight:700; font-size:13px; margin-bottom:8px; display:flex; justify-content:space-between; align-items:center;">
            <span style="color:var(--shopee-color);" id="policy-checker-status-title">🚨 Phát hiện từ khóa rủi ro vi phạm chính sách!</span>
            <button class="copy-btn" style="margin-top:0; padding:4px 8px; font-size:11px;" onclick="cleanTextAndReplace()">Sửa nhanh văn bản</button>
        </div>
        <div id="policy-checker-violations" style="font-size:13px; color: var(--text-primary);">
            <!-- Violated words list -->
        </div>
    </div>
</div>
"""

shopee_campaign_calculator_html = """
<div class="widget-card" id="shopee-campaign-calc-widget">
    <div class="widget-title" style="color: var(--shopee-color); border-bottom: 1px solid var(--border-color); padding-bottom: 8px; margin-bottom: 16px; display: flex; justify-content: space-between; align-items: center;">
        <span style="display: flex; align-items: center; gap: 8px;">
            📊 BẢNG TÍNH LỢI NHUẬN CHIẾN DỊCH SHOPEE (MULTI-ROW GRID)
        </span>
        <a href="TÍNH CHI PHÍ SHOPPE 2.xlsx" class="copy-btn" style="text-decoration: none; padding: 4px 10px; font-size: 11px; margin-top: 0; background: var(--border-color); color: var(--text-primary);" download>
            📥 Tải Excel Gốc
        </a>
    </div>

    <!-- Dashboard Metrics -->
    <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 12px; margin-bottom: 20px;">
        <div style="background: var(--bg-primary); border: 1px solid var(--border-color); padding: 12px; border-radius: 8px; text-align: center;">
            <div style="font-size: 11px; color: var(--text-secondary); margin-bottom: 4px;">Tổng Doanh Thu</div>
            <div id="sp-dash-revenue" style="font-size: 16px; font-weight: 700; color: var(--text-primary);">0 VNĐ</div>
        </div>
        <div style="background: var(--bg-primary); border: 1px solid var(--border-color); padding: 12px; border-radius: 8px; text-align: center;">
            <div style="font-size: 11px; color: var(--text-secondary); margin-bottom: 4px;">Tổng Phí Shopee</div>
            <div id="sp-dash-fees" style="font-size: 16px; font-weight: 700; color: #ef4444;">0 VNĐ</div>
        </div>
        <div style="background: var(--bg-primary); border: 1px solid var(--border-color); padding: 12px; border-radius: 8px; text-align: center;">
            <div style="font-size: 11px; color: var(--text-secondary); margin-bottom: 4px;">Tổng CP Vận Hành</div>
            <div id="sp-dash-cogs" style="font-size: 16px; font-weight: 700; color: var(--text-secondary);">0 VNĐ</div>
        </div>
        <div style="background: var(--bg-primary); border: 1px solid var(--border-color); padding: 12px; border-radius: 8px; text-align: center;">
            <div style="font-size: 11px; color: var(--text-secondary); margin-bottom: 4px;">Lợi Nhuận Ròng</div>
            <div id="sp-dash-profit" style="font-size: 16px; font-weight: 700; color: #22c55e;">0 VNĐ</div>
        </div>
        <div style="background: var(--bg-primary); border: 1px solid var(--border-color); padding: 12px; border-radius: 8px; text-align: center;">
            <div style="font-size: 11px; color: var(--text-secondary); margin-bottom: 4px;">Biên LN Bình Quân</div>
            <div id="sp-dash-margin" style="font-size: 16px; font-weight: 700; color: var(--shopee-color);">0.0%</div>
        </div>
    </div>

    <!-- Scrollable Table -->
    <div style="overflow-x: auto; border: 1px solid var(--border-color); border-radius: 8px; margin-bottom: 16px;">
        <table style="margin: 0; width: 100%; border: none; min-width: 950px;" id="shopee-calc-table">
            <thead>
                <tr>
                    <th style="width: 25%; font-size:12px; padding: 10px;">Sản phẩm / SKU</th>
                    <th style="width: 10%; font-size:12px; padding: 10px;">Loại Shop</th>
                    <th style="width: 11%; font-size:12px; padding: 10px;">Giá Bán (A)</th>
                    <th style="width: 9%; font-size:12px; padding: 10px;">Voucher (C)</th>
                    <th style="width: 8%; font-size:12px; padding: 10px;">Affiliate %</th>
                    <th style="width: 10%; font-size:12px; padding: 10px;">Giá Vốn</th>
                    <th style="width: 10%; font-size:12px; padding: 10px;">CP Khác</th>
                    <th style="width: 12%; font-size:12px; padding: 10px;">Phí Sàn</th>
                    <th style="width: 12%; font-size:12px; padding: 10px;">Lãi Ròng</th>
                    <th style="width: 3%; font-size:12px; padding: 10px; text-align:center;">Xóa</th>
                </tr>
            </thead>
            <tbody id="shopee-calc-tbody">
                <!-- Rows injected dynamically via JS -->
            </tbody>
        </table>
    </div>

    <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 24px;">
        <button class="copy-btn" style="margin-top: 0; background: var(--shopee-color);" onclick="addShopeeCalcRow()">
            ➕ Thêm dòng sản phẩm
        </button>
        <div style="font-size: 12px; color: var(--text-secondary);">
            * Các dòng tự động tính toán độc lập. Tự động lưu trên trình duyệt của bạn.
        </div>
    </div>

    <!-- Product Database Section -->
    <details style="background: var(--bg-primary); border: 1px solid var(--border-color); border-radius: 8px; padding: 12px;">
        <summary style="font-size: 13px; font-weight: 600; cursor: pointer; color: var(--text-secondary); outline:none;">
            ⚙️ Quản lý danh mục sản phẩm Shopee
        </summary>
        <div style="margin-top: 12px;">
            <div style="display: flex; gap: 8px; margin-bottom: 12px;">
                <input type="text" id="shopee-db-search" placeholder="Tìm theo tên/SKU..." onkeyup="searchShopeeDB()" style="flex:1; padding:6px 10px; font-size:12px; background:var(--bg-secondary); border:1px solid var(--border-color); border-radius:4px; color:var(--text-primary); outline:none;">
                <button onclick="openAddShopeeProductModal()" style="padding:6px 12px; font-size:12px; background:var(--shopee-color); color:white; border:none; border-radius:4px; cursor:pointer; font-weight:600;">+ Thêm SKU mới</button>
            </div>
            
            <div style="max-height: 250px; overflow-y: auto; border: 1px solid var(--border-color); border-radius: 4px;">
                <table style="margin: 0; border: none; font-size:11.5px; width:100%;">
                    <thead>
                        <tr>
                            <th style="padding:6px;">SKU</th>
                            <th style="padding:6px;">Tên Sản Phẩm</th>
                            <th style="padding:6px; text-align:right;">% HH Thường</th>
                            <th style="padding:6px; text-align:right;">% HH Mall</th>
                            <th style="padding:6px; text-align:right;">Giá Vốn</th>
                            <th style="padding:6px; text-align:right;">Đóng Gói</th>
                            <th style="padding:6px; text-align:center;">Thao tác</th>
                        </tr>
                    </thead>
                    <tbody id="shopee-db-tbody">
                        <!-- Product rows filled by JS -->
                    </tbody>
                </table>
            </div>

            <div style="display: flex; justify-content: space-between; margin-top: 12px;">
                <button onclick="resetShopeeDB()" style="padding:5px 10px; font-size:11px; background:transparent; border:1px solid var(--border-color); color:var(--text-secondary); border-radius:4px; cursor:pointer;">Đặt lại về mặc định</button>
                <div style="display: flex; gap: 6px;">
                    <button onclick="exportShopeeDB()" style="padding:5px 10px; font-size:11px; background:var(--bg-secondary); border:1px solid var(--border-color); color:var(--text-primary); border-radius:4px; cursor:pointer;">Xuất JSON</button>
                    <button onclick="importShopeeDB()" style="padding:5px 10px; font-size:11px; background:var(--bg-secondary); border:1px solid var(--border-color); color:var(--text-primary); border-radius:4px; cursor:pointer;">Nhập JSON</button>
                </div>
            </div>
        </div>
    </details>
</div>
"""

tiktok_campaign_calculator_html = """
<div class="widget-card" id="tiktok-campaign-calc-widget">
    <div class="widget-title" style="color: var(--tiktok-accent); border-bottom: 1px solid var(--border-color); padding-bottom: 8px; margin-bottom: 16px; display: flex; justify-content: space-between; align-items: center;">
        <span style="display: flex; align-items: center; gap: 8px;">
            📊 BẢNG TÍNH LỢI NHUẬN CHIẾN DỊCH TIKTOK SHOP (MULTI-ROW GRID)
        </span>
        <a href="TÍNH CHI PHÍ TIKTOK.xlsx" class="copy-btn" style="text-decoration: none; padding: 4px 10px; font-size: 11px; margin-top: 0; background: var(--border-color); color: var(--text-primary);" download>
            📥 Tải Excel Gốc
        </a>
    </div>

    <!-- Dashboard Metrics -->
    <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 12px; margin-bottom: 20px;">
        <div style="background: var(--bg-primary); border: 1px solid var(--border-color); padding: 12px; border-radius: 8px; text-align: center;">
            <div style="font-size: 11px; color: var(--text-secondary); margin-bottom: 4px;">Tổng Doanh Thu</div>
            <div id="tt-dash-revenue" style="font-size: 16px; font-weight: 700; color: var(--text-primary);">0 VNĐ</div>
        </div>
        <div style="background: var(--bg-primary); border: 1px solid var(--border-color); padding: 12px; border-radius: 8px; text-align: center;">
            <div style="font-size: 11px; color: var(--text-secondary); margin-bottom: 4px;">Tổng Phí TikTok</div>
            <div id="tt-dash-fees" style="font-size: 16px; font-weight: 700; color: #ef4444;">0 VNĐ</div>
        </div>
        <div style="background: var(--bg-primary); border: 1px solid var(--border-color); padding: 12px; border-radius: 8px; text-align: center;">
            <div style="font-size: 11px; color: var(--text-secondary); margin-bottom: 4px;">Tổng CP Vận Hành</div>
            <div id="tt-dash-cogs" style="font-size: 16px; font-weight: 700; color: var(--text-secondary);">0 VNĐ</div>
        </div>
        <div style="background: var(--bg-primary); border: 1px solid var(--border-color); padding: 12px; border-radius: 8px; text-align: center;">
            <div style="font-size: 11px; color: var(--text-secondary); margin-bottom: 4px;">Lợi Nhuận Ròng</div>
            <div id="tt-dash-profit" style="font-size: 16px; font-weight: 700; color: #22c55e;">0 VNĐ</div>
        </div>
        <div style="background: var(--bg-primary); border: 1px solid var(--border-color); padding: 12px; border-radius: 8px; text-align: center;">
            <div style="font-size: 11px; color: var(--text-secondary); margin-bottom: 4px;">Biên LN Bình Quân</div>
            <div id="tt-dash-margin" style="font-size: 16px; font-weight: 700; color: var(--tiktok-accent);">0.0%</div>
        </div>
    </div>

    <!-- Scrollable Table -->
    <div style="overflow-x: auto; border: 1px solid var(--border-color); border-radius: 8px; margin-bottom: 16px;">
        <table style="margin: 0; width: 100%; border: none; min-width: 950px;" id="tiktok-calc-table">
            <thead>
                <tr>
                    <th style="width: 25%; font-size:12px; padding: 10px;">Sản phẩm / SKU</th>
                    <th style="width: 10%; font-size:12px; padding: 10px;">Loại Shop</th>
                    <th style="width: 11%; font-size:12px; padding: 10px;">Giá Bán (A)</th>
                    <th style="width: 9%; font-size:12px; padding: 10px;">Voucher (C)</th>
                    <th style="width: 8%; font-size:12px; padding: 10px;">Affiliate %</th>
                    <th style="width: 10%; font-size:12px; padding: 10px;">Giá Vốn</th>
                    <th style="width: 10%; font-size:12px; padding: 10px;">CP Khác</th>
                    <th style="width: 12%; font-size:12px; padding: 10px;">Phí Sàn</th>
                    <th style="width: 12%; font-size:12px; padding: 10px;">Lãi Ròng</th>
                    <th style="width: 3%; font-size:12px; padding: 10px; text-align:center;">Xóa</th>
                </tr>
            </thead>
            <tbody id="tiktok-calc-tbody">
                <!-- Rows injected dynamically via JS -->
            </tbody>
        </table>
    </div>

    <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 24px;">
        <button class="copy-btn" style="margin-top: 0; background: var(--tiktok-accent);" onclick="addTikTokCalcRow()">
            ➕ Thêm dòng sản phẩm
        </button>
        <div style="font-size: 12px; color: var(--text-secondary);">
            * Các dòng tự động tính toán độc lập. Tự động lưu trên trình duyệt của bạn.
        </div>
    </div>

    <!-- Product Database Section -->
    <details style="background: var(--bg-primary); border: 1px solid var(--border-color); border-radius: 8px; padding: 12px;">
        <summary style="font-size: 13px; font-weight: 600; cursor: pointer; color: var(--text-secondary); outline:none;">
            ⚙️ Quản lý danh mục sản phẩm TikTok Shop
        </summary>
        <div style="margin-top: 12px;">
            <div style="display: flex; gap: 8px; margin-bottom: 12px;">
                <input type="text" id="tiktok-db-search" placeholder="Tìm theo tên/SKU..." onkeyup="searchTikTokDB()" style="flex:1; padding:6px 10px; font-size:12px; background:var(--bg-secondary); border:1px solid var(--border-color); border-radius:4px; color:var(--text-primary); outline:none;">
                <button onclick="openAddTikTokProductModal()" style="padding:6px 12px; font-size:12px; background:var(--tiktok-accent); color:white; border:none; border-radius:4px; cursor:pointer; font-weight:600;">+ Thêm SKU mới</button>
            </div>
            
            <div style="max-height: 250px; overflow-y: auto; border: 1px solid var(--border-color); border-radius: 4px;">
                <table style="margin: 0; border: none; font-size:11.5px; width:100%;">
                    <thead>
                        <tr>
                            <th style="padding:6px;">SKU</th>
                            <th style="padding:6px;">Tên Sản Phẩm</th>
                            <th style="padding:6px; text-align:right;">% HH Thường</th>
                            <th style="padding:6px; text-align:right;">% HH Mall</th>
                            <th style="padding:6px; text-align:right;">Giá Vốn</th>
                            <th style="padding:6px; text-align:right;">Đóng Gói</th>
                            <th style="padding:6px; text-align:center;">Thao tác</th>
                        </tr>
                    </thead>
                    <tbody id="tiktok-db-tbody">
                        <!-- Product rows filled by JS -->
                    </tbody>
                </table>
            </div>

            <div style="display: flex; justify-content: space-between; margin-top: 12px;">
                <button onclick="resetTikTokDB()" style="padding:5px 10px; font-size:11px; background:transparent; border:1px solid var(--border-color); color:var(--text-secondary); border-radius:4px; cursor:pointer;">Đặt lại về mặc định</button>
                <div style="display: flex; gap: 6px;">
                    <button onclick="exportTikTokDB()" style="padding:5px 10px; font-size:11px; background:var(--bg-secondary); border:1px solid var(--border-color); color:var(--text-primary); border-radius:4px; cursor:pointer;">Xuất JSON</button>
                    <button onclick="importTikTokDB()" style="padding:5px 10px; font-size:11px; background:var(--bg-secondary); border:1px solid var(--border-color); color:var(--text-primary); border-radius:4px; cursor:pointer;">Nhập JSON</button>
                </div>
            </div>
        </div>
    </details>
</div>
"""

# Parser function
def parse_markdown_to_html(content, platform_type="shopee", level=""):
    lines = content.split('\n')
    parsed_html_blocks = []

    in_list = False
    in_ordered_list = False
    in_table = False
    table_rows = []
    in_alert = False
    alert_type = ""
    alert_lines = []
    current_shopee_ch = None

    def parse_inline_markdown(text):
        text = re.sub(r'\*\*(.*?)\*\*', r'<strong>\1</strong>', text)
        text = re.sub(r'`(.*?)`', r'<code>\1</code>', text)
        
        def img_sub(match):
            alt = match.group(1)
            path = match.group(2)
            if "flash_image_" in path:
                new_path = path.replace("images/flash_image_", "shopee_flash_sale_images/image_")
            elif "tiktok_extracted" in path:
                new_path = path
            else:
                new_path = path.replace("images/", "scratch/gdoc_html/images/")
            return f'<img class="handbook-img" src="{new_path}" alt="{alt}">'

        text = re.sub(r'!\[(.*?)\]\((.*?)\)', img_sub, text)
        text = re.sub(r'\[(.*?)\]\((.*?)\)', r'<a href="\2" target="_blank">\1</a>', text)
        return text

    for line in lines:
        line_strip = line.strip()

        if in_alert and not line_strip.startswith('>'):
            alert_content = parse_inline_markdown('<br>'.join(alert_lines))
            title = "💡 Lưu ý"
            if alert_type == "warning":
                title = "⚠️ Cảnh báo quan trọng"
            elif alert_type == "danger" or alert_type == "important":
                title = "🚨 Quy định bắt buộc"
            parsed_html_blocks.append(f'<div class="alert-box alert-{alert_type}"><div class="alert-title">{title}</div><p>{alert_content}</p></div>')
            in_alert = False
            alert_lines = []

        if in_table and not line_strip.startswith('|'):
            clean_rows = []
            for r in table_rows:
                if not re.match(r'^\s*\|\s*:?-+:?\s*\|', r):
                    clean_rows.append(r)
            
            table_html = '<table>'
            for idx, row in enumerate(clean_rows):
                cols = [c.strip() for c in row.split('|')[1:-1]]
                table_html += '<tr>'
                for col in cols:
                    cell_content = parse_inline_markdown(col)
                    if idx == 0:
                        table_html += f'<th>{cell_content}</th>'
                    else:
                        table_html += f'<td>{cell_content}</td>'
                table_html += '</tr>'
            table_html += '</table>'
            parsed_html_blocks.append(table_html)
            in_table = False
            table_rows = []

        if in_list and not (line_strip.startswith('*') or line_strip.startswith('-') or line_strip.startswith('    *') or line_strip.startswith('    -') or not line_strip):
            if line_strip:
                parsed_html_blocks.append('</ul>')
                in_list = False

        if in_ordered_list and not (re.match(r'^\d+\.', line_strip) or line_strip.startswith('    ') or not line_strip):
            if line_strip:
                parsed_html_blocks.append('</ol>')
                in_ordered_list = False

        if not line_strip:
            continue

        if line_strip.startswith('# ') or line_strip.startswith('---'):
            continue
        
        # TikTok sections
        if platform_type == "tiktok":
            suffix = f"-{level}" if level else ""
            mod_match = re.match(r'^##\s*Module\s*(\d+)\s*-\s*(.*)', line_strip, re.IGNORECASE)
            if mod_match:
                mod_num = mod_match.group(1)
                mod_title = mod_match.group(2).strip()
                parsed_html_blocks.append(f'<h1 id="tiktok-chuong{mod_num}{suffix}" class="handbook-section-h1 tiktok-section-{level}">Module {mod_num}: {mod_title}</h1>')
                
                # Inject TikTok Onboarding checklist under Module 1 (for both)
                if mod_num == "1":
                    tiktok_checklist_html = f"""
                    <div class="checklist-container" id="tiktok-onboarding-checklist-widget-{level}">
                        <div class="checklist-header">
                            <div style="font-weight: 700; color: var(--accent-color); font-size: 15px; display: flex; align-items: center; gap: 8px;">
                                🎯 BẢNG TIẾN TRÌNH THỰC HÀNH MỞ SHOP TIKTOK
                            </div>
                            <div id="tiktok-checklist-pct-{level}" style="font-weight: 700; font-size: 13px;">Tiến độ: 0%</div>
                        </div>
                        <div class="checklist-progress-bar-bg">
                            <div class="checklist-progress-bar-fill" id="tiktok-checklist-progress-fill-{level}"></div>
                        </div>
                        <div id="tiktok-checklist-items-{level}">
                            <label class="checklist-item">
                                <input type="checkbox" id="tt-chk-{level}-0" onchange="toggleTikTokChecklistItem('{level}', 0)">
                                <span class="checklist-item-text"><strong>Bước 1:</strong> Thiết lập tài khoản TikTok mới, cập nhật ảnh đại diện & viết mô tả Bio chuẩn định vị.</span>
                            </label>
                            <label class="checklist-item">
                                <input type="checkbox" id="tt-chk-{level}-1" onchange="toggleTikTokChecklistItem('{level}', 1)">
                                <span class="checklist-item-text"><strong>Bước 2:</strong> Lên lịch sản xuất & đăng tối thiểu 7-10 video content nền giá trị để tạo độ tin cậy.</span>
                            </label>
                            <label class="checklist-item">
                                <input type="checkbox" id="tt-chk-{level}-2" onchange="toggleTikTokChecklistItem('{level}', 2)">
                                <span class="checklist-item-text"><strong>Bước 3:</strong> Đăng ký tài khoản Seller Center (Sử dụng NFC xác thực CCCD gắn chip).</span>
                            </label>
                            <label class="checklist-item">
                                <input type="checkbox" id="tt-chk-{level}-3" onchange="toggleTikTokChecklistItem('{level}', 3)">
                                <span class="checklist-item-text"><strong>Bước 4:</strong> Thiết lập chính xác địa chỉ kho lấy hàng & kho hoàn trả hàng.</span>
                            </label>
                            <label class="checklist-item">
                                <input type="checkbox" id="tt-chk-{level}-4" onchange="toggleTikTokChecklistItem('{level}', 4)">
                                <span class="checklist-item-text"><strong>Bước 5:</strong> Liên kết Tài khoản TikTok chính thức (Official Account) để đồng bộ giỏ hàng lên hồ sơ.</span>
                            </label>
                            <label class="checklist-item">
                                <input type="checkbox" id="tt-chk-{level}-5" onchange="toggleTikTokChecklistItem('{level}', 5)">
                                <span class="checklist-item-text"><strong>Bước 6:</strong> Đăng sản phẩm đầu tiên, tính toán cơ cấu chi phí sàn & thiết lập giá bán có lãi.</span>
                            </label>
                        </div>
                    </div>
                    """
                    parsed_html_blocks.append(tiktok_checklist_html)
                
                # Inject campaign profit calculator under Module 10 (Advanced only)
                if mod_num == "10" and level == "advanced":
                    parsed_html_blocks.append(tiktok_campaign_calculator_html)
                
                continue

            part_match = re.match(r'^##\s*Phần\s*(\d+)\s*-\s*(.*)', line_strip, re.IGNORECASE)
            if part_match:
                part_num = part_match.group(1)
                part_title = part_match.group(2).strip()
                parsed_html_blocks.append(f'<h1 id="tiktok-chuong{part_num}{suffix}" class="handbook-section-h1 tiktok-section-{level}">Phần {part_num}: {part_title}</h1>')
                continue

            if line_strip.startswith("## Phụ lục"):
                clean = re.sub(r'^##\s*', '', line_strip)
                parsed_html_blocks.append(f'<h1 id="tiktok-chuong19{suffix}" class="handbook-section-h1 tiktok-section-{level}">{clean}</h1>')
                continue

            if "Mục lục" in line_strip:
                clean = re.sub(r'^##\s*', '', line_strip)
                parsed_html_blocks.append(f'<h2 class="handbook-section-h2">{clean}</h2>')
                continue

        # Shopee chapters
        ch_match = re.match(r'^##\s*CHƯƠNG\s*(\d+):\s*(.*)', line_strip, re.IGNORECASE)
        if ch_match:
            ch_num = ch_match.group(1)
            current_shopee_ch = ch_num
            ch_title = ch_match.group(2).strip()
            parsed_html_blocks.append(f'<h1 id="h1-chuong{ch_num}" class="handbook-section-h1 shopee-section">CHƯƠNG {ch_num}: {ch_title}</h1>')
            
            # Inject Shopee Onboarding Checklist under Shopee Chapter 1
            if ch_num == "1":
                onboarding_checklist_html = """
                <div class="checklist-container" id="onboarding-checklist-widget">
                    <div class="checklist-header">
                        <div style="font-weight: 700; color: var(--accent-color); font-size: 15px; display: flex; align-items: center; gap: 8px;">
                            🎯 BẢNG TIẾN TRÌNH THỰC HÀNH MỞ SHOP (Dành cho Người mới)
                        </div>
                        <div id="checklist-pct" style="font-weight: 700; font-size: 13px;">Tiến độ: 0%</div>
                    </div>
                    <div class="checklist-progress-bar-bg">
                        <div class="checklist-progress-bar-fill" id="checklist-progress-fill"></div>
                    </div>
                    <div id="checklist-items">
                        <label class="checklist-item">
                            <input type="checkbox" id="chk-0" onchange="toggleChecklistItem(0)">
                            <span class="checklist-item-text"><strong>Bước 1:</strong> Đăng ký tài khoản Shopee mới & Đăng nhập thành công vào Kênh Người Bán</span>
                        </label>
                        <label class="checklist-item">
                            <input type="checkbox" id="chk-1" onchange="toggleChecklistItem(1)">
                            <span class="checklist-item-text"><strong>Bước 2:</strong> Thiết lập Thông tin Shop, địa chỉ kho lấy hàng & số điện thoại bưu tá liên hệ</span>
                        </label>
                        <label class="checklist-item">
                            <input type="checkbox" id="chk-2" onchange="toggleChecklistItem(2)">
                            <span class="checklist-item-text"><strong>Bước 3:</strong> Kích hoạt các kênh vận chuyển phù hợp (Hỏa Tốc, Nhanh, Hàng Cồng Kềnh)</span>
                        </label>
                        <label class="checklist-item">
                            <input type="checkbox" id="chk-3" onchange="toggleChecklistItem(3)">
                            <span class="checklist-item-text"><strong>Bước 4:</strong> Cập nhật Mã số thuế (MST) & Gửi hình ảnh CCCD để định danh KYC tài khoản</span>
                        </label>
                        <label class="checklist-item">
                            <input type="checkbox" id="chk-4" onchange="toggleChecklistItem(4)">
                            <span class="checklist-item-text"><strong>Bước 5:</strong> Liên kết Tài khoản ngân hàng rút tiền trùng khớp 100% với định danh KYC (Luật 122)</span>
                        </label>
                        <label class="checklist-item">
                            <input type="checkbox" id="chk-5" onchange="toggleChecklistItem(5)">
                            <span class="checklist-item-text"><strong>Bước 6:</strong> Sử dụng công cụ Trang trí Shop bản điện thoại để thiết kế giao diện gian hàng</span>
                        </label>
                    </div>
                </div>
                """
                parsed_html_blocks.append(onboarding_checklist_html)
            continue

        if line_strip.startswith('## '):
            clean = re.sub(r'^##\s*', '', line_strip)
            parsed_html_blocks.append(f'<h2 class="handbook-section-h2">{parse_inline_markdown(clean)}</h2>')
            continue

        if line_strip.startswith('### '):
            clean = re.sub(r'^###\s*', '', line_strip)
            if platform_type == "shopee" and current_shopee_ch is not None:
                sec_match = re.match(r'^(\d+)\.\s*(.*)', clean)
                if sec_match:
                    sec_num = sec_match.group(1)
                else:
                    sec_num = "unknown"
                parsed_html_blocks.append(f'<h3 id="shopee-ch{current_shopee_ch}-sec{sec_num}" class="handbook-section-h3 shopee-sec-header">{parse_inline_markdown(clean)}</h3>')
            else:
                parsed_html_blocks.append(f'<h3 class="handbook-section-h3">{parse_inline_markdown(clean)}</h3>')
            
            # Inject OBS Configurator for Shopee (Chapter 3)
            if platform_type == "shopee" and "OBS Livestream" in clean:
                parsed_html_blocks.append(obs_configurator_html)
            
            # Inject Return/Refund Dispute Assistant for Shopee (Chapter 5)
            if platform_type == "shopee" and "Khiếu nại quyết định" in clean:
                parsed_html_blocks.append(dispute_assistant_html)

            # Inject TikTok Policy checker widget (TikTok Module 8)
            if platform_type == "tiktok" and ("Từ khóa an toàn chính sách" in line_strip or "Chính sách ngành hàng" in clean):
                parsed_html_blocks.append(tiktok_policy_checker_html)
            continue

        if line_strip.startswith('#### '):
            clean = re.sub(r'^####\s*', '', line_strip)
            parsed_html_blocks.append(f'<h4 class="handbook-section-h4">{parse_inline_markdown(clean)}</h4>')
            continue

        if line_strip.startswith('> '):
            alert_content = re.sub(r'^>\s*', '', line_strip)
            if alert_content.startswith('[!NOTE]'):
                in_alert = True
                alert_type = "info"
                continue
            elif alert_content.startswith('[!WARNING]'):
                in_alert = True
                alert_type = "warning"
                continue
            elif alert_content.startswith('[!IMPORTANT]') or alert_content.startswith('[!CAUTION]'):
                in_alert = True
                alert_type = "danger"
                continue
            
            if in_alert:
                alert_lines.append(alert_content)
            else:
                parsed_html_blocks.append(f'<blockquote>{parse_inline_markdown(alert_content)}</blockquote>')
            continue

        if line_strip.startswith('|'):
            in_table = True
            table_rows.append(line_strip)
            continue

        if line_strip.startswith('* ') or line_strip.startswith('- ') or line_strip.startswith('    * ') or line_strip.startswith('    - '):
            clean = re.sub(r'^(\s*[\*\-]\s*)', '', line)
            if not in_list:
                parsed_html_blocks.append('<ul>')
                in_list = True
            parsed_html_blocks.append(f'<li>{parse_inline_markdown(clean)}</li>')
            continue

        num_match = re.match(r'^(\d+)\.\s*(.*)', line_strip)
        if num_match:
            clean = num_match.group(2)
            if not in_ordered_list:
                parsed_html_blocks.append('<ol>')
                in_ordered_list = True
            parsed_html_blocks.append(f'<li>{parse_inline_markdown(clean)}</li>')
            continue

        if line_strip.startswith('$$') and line_strip.endswith('$$'):
            clean = line_strip.strip('$').strip()
            parsed_html_blocks.append(f'<div class="math-block">{clean}</div>')
            continue

        parsed_html_blocks.append(f'<p>{parse_inline_markdown(line_strip)}</p>')

    # Close left blocks
    if in_list:
        parsed_html_blocks.append('</ul>')
    if in_ordered_list:
        parsed_html_blocks.append('</ol>')
    if in_alert:
        alert_content = parse_inline_markdown('<br>'.join(alert_lines))
        parsed_html_blocks.append(f'<div class="alert-box alert-{alert_type}"><div class="alert-title">💡 Lưu ý</div><p>{alert_content}</p></div>')

    return '\n'.join(parsed_html_blocks)

# Helper to compile Shopee chapter submenus
def compile_shopee_chapter_menu(ch_num, ch_title, subsections):
    if not subsections:
        return f'<li class="menu-item" id="menu-item-shopee-ch{ch_num}"><a class="menu-link" onclick="scrollToSection(\'h1-chuong{ch_num}\')">Chương {ch_num}: {ch_title}</a></li>'
    
    sub_html = []
    for sec_num, sec_title in subsections:
        sub_html.append(f'<li class="submenu-item"><a class="submenu-link shopee-sublink" id="sublink-shopee-ch{ch_num}-sec{sec_num}" onclick="scrollToSection(\'shopee-ch{ch_num}-sec{sec_num}\')">{sec_num}. {sec_title}</a></li>')
        
    sub_list_str = '\n'.join(sub_html)
    
    return f"""<li class="menu-item has-submenu" id="menu-item-shopee-ch{ch_num}">
        <div class="menu-header-wrapper" onclick="toggleSidebarSubmenu('shopee-ch{ch_num}'); scrollToSection('h1-chuong{ch_num}')">
            <a class="menu-link">Chương {ch_num}: {ch_title}</a>
            <span class="submenu-arrow">▶</span>
        </div>
        <ul class="submenu-list" id="submenu-shopee-ch{ch_num}" style="display: none;">
            {sub_list_str}
        </ul>
    </li>"""

# Generate sidebar menu items dynamically from markdown files
def generate_sidebar_menu(content, platform_type="shopee", level=""):
    lines = content.split('\n')
    menu_html = []
    
    if platform_type == "shopee":
        current_ch = None
        current_ch_title = ""
        current_subsections = []
        
        for line in lines:
            line_strip = line.strip()
            
            # Check for Chapter (H2 with "CHƯƠNG")
            ch_match = re.match(r'^##\s*CHƯƠNG\s*(\d+):\s*(.*)', line_strip, re.IGNORECASE)
            if ch_match:
                if current_ch is not None:
                    menu_html.append(compile_shopee_chapter_menu(current_ch, current_ch_title, current_subsections))
                current_ch = ch_match.group(1)
                current_ch_title = ch_match.group(2).strip()
                current_subsections = []
                continue
            
            # Check for Subsection (H3)
            if current_ch is not None and line_strip.startswith('### '):
                clean_sec = re.sub(r'^###\s*', '', line_strip).strip()
                sec_match = re.match(r'^(\d+)\.\s*(.*)', clean_sec)
                if sec_match:
                    sec_num = sec_match.group(1)
                    sec_title = sec_match.group(2).strip()
                else:
                    sec_num = str(len(current_subsections) + 1)
                    sec_title = clean_sec
                sec_title_clean = re.sub(r'\*\*|`|_', '', sec_title)
                current_subsections.append((sec_num, sec_title_clean))
        
        if current_ch is not None:
            menu_html.append(compile_shopee_chapter_menu(current_ch, current_ch_title, current_subsections))
            
        return '\n'.join(menu_html)
        
    else: # tiktok
        for line in lines:
            line_strip = line.strip()
            if not line_strip.startswith('## '):
                continue
            suffix = f"-{level}" if level else ""
            mod_match = re.match(r'^##\s*Module\s*(\d+)\s*-\s*(.*)', line_strip, re.IGNORECASE)
            if mod_match:
                mod_num = mod_match.group(1)
                mod_title = mod_match.group(2).strip()
                menu_html.append(f"""<li class="menu-item" id="menu-item-tiktok-{level}-mod{mod_num}">
                    <div class="menu-header-wrapper" onclick="scrollToSection('tiktok-chuong{mod_num}{suffix}')">
                        <a class="menu-link">M{mod_num}: {mod_title}</a>
                    </div>
                </li>""")
                continue
                
            part_match = re.match(r'^##\s*Phần\s*(\d+)\s*-\s*(.*)', line_strip, re.IGNORECASE)
            if part_match:
                part_num = part_match.group(1)
                part_title = part_match.group(2).strip()
                menu_html.append(f"""<li class="menu-item" id="menu-item-tiktok-{level}-part{part_num}">
                    <div class="menu-header-wrapper" onclick="scrollToSection('tiktok-chuong{part_num}{suffix}')">
                        <a class="menu-link">Phần {part_num}: {part_title}</a>
                    </div>
                </li>""")
                continue
                
            if "Phụ lục" in line_strip:
                clean = re.sub(r'^##\s*', '', line_strip)
                menu_html.append(f"""<li class="menu-item" id="menu-item-tiktok-{level}-appendix">
                    <div class="menu-header-wrapper" onclick="scrollToSection('tiktok-chuong19{suffix}')">
                        <a class="menu-link">{clean}</a>
                    </div>
                </li>""")
                continue
        return '\n'.join(menu_html)

shopee_menu_items = generate_sidebar_menu(shopee_content, "shopee")
tiktok_basic_menu_items = generate_sidebar_menu(tiktok_basic_content, "tiktok", "basic")
tiktok_advanced_menu_items = generate_sidebar_menu(tiktok_advanced_content, "tiktok", "advanced")

# Parse Markdowns
shopee_html = parse_markdown_to_html(shopee_content, "shopee")

# Add Shopee Tax Calculator
shopee_tax_calculator_html = """
<div class="widget-card" id="tax-calculator-widget">
    <div class="widget-title">
        <svg viewBox="0 0 24 24" width="20" height="20" fill="currentColor">
            <path d="M19 3H5c-1.1 0-2 .9-2 2v14c0 1.1.9 2 2 2h14c1.1 0 2-.9 2-2V5c0-1.1-.9-2-2-2zm-4 6h-4v2h4v2h-4v2h4v2H9V7h6v2z"/>
        </svg>
        Công cụ tính Thuế sàn Shopee khấu trừ nộp thay (NĐ 117/2025)
    </div>
    <p style="font-size: 13px; color: var(--text-secondary); margin-bottom: 16px;">
        Công thức áp dụng theo NĐ 117/2025/NĐ-CP:<br>
        <code>Doanh thu tính thuế (G) = A - C - D2</code><br>
        <code>Thuế khấu trừ = G x (Thuế suất GTGT + Thuế suất TNCN)</code>
    </p>
    <div class="calculator-grid">
        <div>
            <div class="form-group">
                <label for="calc-price">Giá trị hàng hóa niêm yết (A) - VNĐ:</label>
                <input type="number" id="calc-price" value="500000" oninput="calculateTax()">
            </div>
            <div class="form-group">
                <label for="calc-voucher-seller">Mã giảm giá Người bán tự tạo (C) - VNĐ:</label>
                <input type="number" id="calc-voucher-seller" value="50000" oninput="calculateTax()">
            </div>
            <div class="form-group">
                <label for="calc-voucher-co">Mã giảm giá Đồng tài trợ do Shop chịu (D2) - VNĐ:</label>
                <input type="number" id="calc-voucher-co" value="10000" oninput="calculateTax()">
            </div>
            <div class="form-group">
                <label for="calc-category">Ngành hàng / Thuế suất áp dụng:</label>
                <select id="calc-category" onchange="calculateTax()">
                    <option value="goods">Hàng hóa (1.5% = 1.0% GTGT + 0.5% TNCN)</option>
                    <option value="services">Dịch vụ / e-Voucher (7.0% = 5.0% GTGT + 2.0% TNCN)</option>
                </select>
            </div>
        </div>
        <div class="calc-result-box">
            <div class="result-row">
                <span>Doanh thu tính thuế (G):</span>
                <span id="res-taxable-revenue" style="font-weight:600;">440,000 VNĐ</span>
            </div>
            <div class="result-row">
                <span>Thuế Giá trị gia tăng (GTGT):</span>
                <span id="res-vat" style="font-weight:600; color:#3b82f6;">4,400 VNĐ</span>
            </div>
            <div class="result-row">
                <span>Thuế Thu nhập cá nhân (TNCN):</span>
                <span id="res-pit" style="font-weight:600; color:#3b82f6;">2,200 VNĐ</span>
            </div>
            <div class="result-row total">
                <span>Tổng thuế Shopee khấu trừ:</span>
                <span id="res-total-tax">6,600 VNĐ</span>
            </div>
        </div>
    </div>
</div>
"""

shopee_html += shopee_campaign_calculator_html
shopee_html += shopee_tax_calculator_html

tiktok_basic_html = parse_markdown_to_html(tiktok_basic_content, "tiktok", "basic")
tiktok_advanced_html = parse_markdown_to_html(tiktok_advanced_content, "tiktok", "advanced")

# Body Content Structure with dual Basic vs Advanced versions
body_content = f"""
    <!-- Sidebar -->
    <aside class="sidebar">
        <div class="sidebar-header">
            <!-- Platform Switcher Tabs -->
            <div class="platform-tabs">
                <button class="platform-tab-btn active" id="tab-shopee" onclick="switchPlatform('shopee')">
                    🧡 Shopee
                </button>
                <button class="platform-tab-btn" id="tab-tiktok" onclick="switchPlatform('tiktok')">
                    🖤 TikTok Shop
                </button>
            </div>
            
            <!-- TikTok Level Switcher (Low profile) -->
            <div id="tiktok-level-selector" class="level-tabs" style="display: none; margin-bottom: 12px;">
                <button class="level-tab-btn active" id="btn-tiktok-basic" onclick="switchTikTokLevel('basic')">
                    Bản Cơ Bản
                </button>
                <button class="level-tab-btn" id="btn-tiktok-advanced" onclick="switchTikTokLevel('advanced')">
                    Bản Nâng Cao
                </button>
            </div>
            
            <div class="search-box">
                <svg fill="none" stroke="currentColor" stroke-linecap="round" stroke-linejoin="round" stroke-width="2" viewBox="0 0 24 24">
                    <circle cx="11" cy="11" r="8"></circle>
                    <line x1="21" y1="21" x2="16.65" y2="16.65"></line>
                </svg>
                <input type="text" id="search-input" placeholder="Tìm kiếm nội dung..." onkeyup="searchContent()">
            </div>
            
            <button class="theme-toggle-btn" onclick="toggleTheme()">
                <svg id="theme-icon" viewBox="0 0 24 24" width="18" height="18" fill="currentColor">
                    <path d="M12 3a9 9 0 1 0 9 9c0-.46-.04-.92-.1-1.36a5.389 5.389 0 0 1-4.4 2.26 5.403 5.403 0 0 1-3.14-9.8c.44-.06.9-.1 1.36-.1H12z"/>
                </svg>
                <span id="theme-text">Chế độ Sáng</span>
            </button>
        </div>

        <nav class="sidebar-menu">
            <!-- Shopee Sidebar Menu -->
            <div id="shopee-menu" class="menu-container">
                <div class="menu-title">Lộ trình Shopee</div>
                <ul class="menu-list">
                    {shopee_menu_items}
                </ul>
            </div>

            <!-- TikTok Shop Basic Sidebar Menu -->
            <div id="tiktok-basic-menu" class="menu-container" style="display:none;">
                <div class="menu-title">TikTok Shop - Cơ Bản</div>
                <ul class="menu-list">
                    {tiktok_basic_menu_items}
                </ul>
            </div>

            <!-- TikTok Shop Advanced Sidebar Menu -->
            <div id="tiktok-advanced-menu" class="menu-container" style="display:none;">
                <div class="menu-title">TikTok Shop - Nâng Cao</div>
                <ul class="menu-list">
                    {tiktok_advanced_menu_items}
                </ul>
            </div>
        </nav>
    </aside>

    <!-- Main Content Area -->
    <main class="content-container" id="content-area">
        <div class="content-header">
            <h3 id="app-subtitle" style="margin-top:0; color:var(--text-secondary); font-size: 14px; font-weight:500;">Cẩm nang Vận hành Bán hàng Shopee Việt Nam v2026</h3>
            <span style="font-size:12px; color:var(--text-secondary); background:var(--border-color); padding: 4px 10px; border-radius: 12px; font-weight:500;">Tài liệu Đào tạo Nội bộ</span>
        </div>
        <div class="content-body">
            <!-- Shopee Section Content -->
            <div id="shopee-content" class="platform-content active">
                <div style="text-align: center; margin-bottom: 40px; padding: 40px; background: var(--bg-secondary); border-radius: 16px; border: 1px solid var(--border-color);">
                    <span style="font-size: 40px;">🍊</span>
                    <h1 style="margin-top: 16px; border:none; padding:0; font-size:32px; line-height:1.2; color:var(--shopee-color);">CẨM NANG VẬN HÀNH SHOPEE</h1>
                    <p style="color: var(--text-secondary); font-size: 16px; margin-top: 12px; margin-bottom:0;">Chính sách Pháp lý - Kỹ thuật Livestream - Hướng dẫn Thuế mới nhất</p>
                </div>
                {shopee_html}
            </div>

            <!-- TikTok Basic Section Content -->
            <div id="tiktok-basic-content" class="platform-content">
                <div style="text-align: center; margin-bottom: 40px; padding: 40px; background: var(--bg-secondary); border-radius: 16px; border: 1px solid var(--border-color);">
                    <span style="font-size: 40px;">🎬</span>
                    <h1 style="margin-top: 16px; border:none; padding:0; font-size:32px; line-height:1.2; color:var(--tiktok-accent);">CẨM NANG TIKTOK SHOP (CƠ BẢN)</h1>
                    <p style="color: var(--text-secondary); font-size: 16px; margin-top: 12px; margin-bottom:0;">Lộ trình làm quen - Đăng ký shop cá nhân - Thiết lập kho & vận hành đơn lẻ</p>
                </div>
                {tiktok_basic_html}
            </div>

            <!-- TikTok Advanced Section Content -->
            <div id="tiktok-advanced-content" class="platform-content">
                <div style="text-align: center; margin-bottom: 40px; padding: 40px; background: var(--bg-secondary); border-radius: 16px; border: 1px solid var(--border-color);">
                    <span style="font-size: 40px;">⚡</span>
                    <h1 style="margin-top: 16px; border:none; padding:0; font-size:32px; line-height:1.2; color:var(--tiktok-accent);">CẨM NANG TIKTOK SHOP (NÂNG CAO)</h1>
                    <p style="color: var(--text-secondary); font-size: 16px; margin-top: 12px; margin-bottom:0;">Chuyên sâu 18 Module - Kỹ thuật Live chuyên nghiệp - Đối soát & Phí sàn chi tiết</p>
                </div>
                {tiktok_advanced_html}
            </div>
"""

# Modals markup for adding new products
modals_html = """
    <!-- Shopee Add SKU Modal -->
    <div id="shopee-add-modal" class="modal-overlay">
        <div class="modal-card">
            <div class="modal-header">
                <span>➕ Thêm sản phẩm Shopee mới</span>
                <button class="modal-close-btn" onclick="closeShopeeAddModal()">&times;</button>
            </div>
            <div class="form-group">
                <label>Mã SKU:</label>
                <input type="text" id="add-sp-sku" placeholder="Ví dụ: HC999">
            </div>
            <div class="form-group">
                <label>Tên sản phẩm:</label>
                <input type="text" id="add-sp-name" placeholder="Ví dụ: Metacare Opti 3+">
            </div>
            <div class="form-group">
                <label>Giá bán tham khảo (VNĐ):</label>
                <input type="number" id="add-sp-price" value="300000">
            </div>
            <div class="form-group">
                <label>% Hoa hồng Shop thường:</label>
                <input type="number" id="add-sp-comm-regular" value="11" step="0.1">
            </div>
            <div class="form-group">
                <label>% Hoa hồng Shop Mall:</label>
                <input type="number" id="add-sp-comm-mall" value="12.5" step="0.1">
            </div>
            <div class="form-group">
                <label>Giá vốn sản phẩm (VNĐ):</label>
                <input type="number" id="add-sp-cogs" value="200000">
            </div>
            <div class="form-group">
                <label>Chi phí đóng gói (VNĐ):</label>
                <input type="number" id="add-sp-packing" value="6000">
            </div>
            <div class="modal-footer">
                <button class="btn-secondary" onclick="closeShopeeAddModal()">Hủy</button>
                <button class="btn-primary" style="background:var(--shopee-color);" onclick="saveCustomShopeeProduct()">Lưu sản phẩm</button>
            </div>
        </div>
    </div>

    <!-- TikTok Add SKU Modal -->
    <div id="tiktok-add-modal" class="modal-overlay">
        <div class="modal-card">
            <div class="modal-header">
                <span>➕ Thêm sản phẩm TikTok mới</span>
                <button class="modal-close-btn" onclick="closeTikTokAddModal()">&times;</button>
            </div>
            <div class="form-group">
                <label>Mã SKU:</label>
                <input type="text" id="add-tt-sku" placeholder="Ví dụ: HC999">
            </div>
            <div class="form-group">
                <label>Tên sản phẩm:</label>
                <input type="text" id="add-tt-name" placeholder="Ví dụ: Metacare Opti 3+">
            </div>
            <div class="form-group">
                <label>% Hoa hồng Shop thường:</label>
                <input type="number" id="add-tt-comm-regular" value="10" step="0.1">
            </div>
            <div class="form-group">
                <label>% Hoa hồng Shop Mall:</label>
                <input type="number" id="add-tt-comm-mall" value="12.5" step="0.1">
            </div>
            <div class="form-group">
                <label>Giá vốn sản phẩm (VNĐ):</label>
                <input type="number" id="add-tt-cogs" value="200000">
            </div>
            <div class="form-group">
                <label>Chi phí đóng gói (VNĐ):</label>
                <input type="number" id="add-tt-packing" value="5000">
            </div>
            <div class="modal-footer">
                <button class="btn-secondary" onclick="closeTikTokAddModal()">Hủy</button>
                <button class="btn-primary" style="background:var(--tiktok-accent);" onclick="saveCustomTikTokProduct()">Lưu sản phẩm</button>
            </div>
        </div>
    </div>
"""

# Script Templates
template_foot = f"""
        </div>
    </main>

    <!-- Lightbox Modal -->
    <div id="lightbox" class="lightbox-modal" onclick="closeLightbox()">
        <span class="lightbox-close">&times;</span>
        <img class="lightbox-content" id="lightbox-img">
        <div class="lightbox-caption" id="lightbox-caption"></div>
    </div>

    {modals_html}

    <script>
        // ----------------------------------------------------
        // 1. Data Definitions (Injected by Builder)
        // ----------------------------------------------------
        const DEFAULT_SHOPEE_PRODUCTS = {json.dumps(shopee_products_db, ensure_ascii=False)};
        const DEFAULT_TIKTOK_PRODUCTS = {json.dumps(tiktok_products_db, ensure_ascii=False)};

        let shopeeProducts = [];
        let tiktokProducts = [];

        let shopeeCalcRows = [];
        let tiktokCalcRows = [];

        let activeTikTokLevel = 'basic'; // Default level

        // ----------------------------------------------------
        // 2. LocalStorage Sync & Init
        // ----------------------------------------------------
        function initProductDatabases() {{
            // Shopee DB
            const savedShopeeDB = localStorage.getItem('shopee_product_db');
            if (savedShopeeDB) {{
                try {{
                    shopeeProducts = JSON.parse(savedShopeeDB);
                }} catch(e) {{
                    shopeeProducts = [...DEFAULT_SHOPEE_PRODUCTS];
                }}
            }} else {{
                shopeeProducts = [...DEFAULT_SHOPEE_PRODUCTS];
                localStorage.setItem('shopee_product_db', JSON.stringify(shopeeProducts));
            }}

            // TikTok DB
            const savedTikTokDB = localStorage.getItem('tiktok_product_db');
            if (savedTikTokDB) {{
                try {{
                    tiktokProducts = JSON.parse(savedTikTokDB);
                }} catch(e) {{
                    tiktokProducts = [...DEFAULT_TIKTOK_PRODUCTS];
                }}
            }} else {{
                tiktokProducts = [...DEFAULT_TIKTOK_PRODUCTS];
                localStorage.setItem('tiktok_product_db', JSON.stringify(tiktokProducts));
            }}

            // Load Calculator states
            const savedShopeeRows = localStorage.getItem('shopee_calc_rows');
            if (savedShopeeRows) {{
                try {{
                    shopeeCalcRows = JSON.parse(savedShopeeRows);
                }} catch(e) {{}}
            }}
            if (shopeeCalcRows.length === 0 && shopeeProducts.length > 0) {{
                shopeeCalcRows.push({{
                    sku: shopeeProducts[0].sku,
                    shopType: 'regular',
                    price: shopeeProducts[0].ref_price || 200000,
                    voucher: 0,
                    affiliateRate: 0,
                    cogs: null,
                    ads: 0,
                    ship: 0
                }});
            }}

            const savedTikTokRows = localStorage.getItem('tiktok_calc_rows');
            if (savedTikTokRows) {{
                try {{
                    tiktokCalcRows = JSON.parse(savedTikTokRows);
                }} catch(e) {{}}
            }}
            if (tiktokCalcRows.length === 0 && tiktokProducts.length > 0) {{
                tiktokCalcRows.push({{
                    sku: tiktokProducts[0].sku,
                    shopType: 'regular',
                    price: 200000,
                    voucher: 0,
                    affiliateRate: 0,
                    cogs: null,
                    ads: 0,
                    ship: 0
                }});
            }}

            // Load TikTok active level
            const savedLevel = localStorage.getItem('tiktok_active_level');
            if (savedLevel) {{
                activeTikTokLevel = savedLevel;
            }}
        }}

        initProductDatabases();

        // ----------------------------------------------------
        // 3. Platform Switching Logic
        // ----------------------------------------------------
        function switchPlatform(platform) {{
            document.documentElement.setAttribute('data-platform', platform);
            
            // Hide all platform contents
            document.querySelectorAll('.platform-content').forEach(div => {{
                div.classList.remove('active');
            }});
            
            // Hide level selector
            document.getElementById('tiktok-level-selector').style.display = 'none';
            
            // Hide all menus
            document.querySelectorAll('.menu-container').forEach(menu => {{
                menu.style.display = 'none';
            }});
            
            // Reset tab active states
            document.querySelectorAll('.platform-tab-btn').forEach(btn => {{
                btn.classList.remove('active');
            }});
            document.getElementById('tab-' + platform).classList.add('active');

            if (platform === 'shopee') {{
                document.getElementById('shopee-content').classList.add('active');
                document.getElementById('shopee-menu').style.display = 'block';
                document.getElementById('app-subtitle').innerText = "Cẩm nang Vận hành Bán hàng Shopee Việt Nam v2026";
                
                // Highlight first chapter
                const activeLink = document.querySelector('#shopee-menu .menu-link');
                document.querySelectorAll('.menu-link').forEach(link => link.classList.remove('active'));
                if (activeLink) activeLink.classList.add('active');
            }} else {{
                // TikTok active: show level selector & appropriate level content
                document.getElementById('tiktok-level-selector').style.display = 'flex';
                switchTikTokLevel(activeTikTokLevel);
            }}
            
            // Scroll content area back to top
            document.getElementById('content-area').scrollTop = 0;
            
            // Clear search field
            document.getElementById('search-input').value = '';
            searchContent();
        }}

        // ----------------------------------------------------
        // 4. TikTok Level Switching (Basic vs Advanced)
        // ----------------------------------------------------
        function switchTikTokLevel(level) {{
            activeTikTokLevel = level;
            localStorage.setItem('tiktok_active_level', level);
            
            // Toggle active buttons style
            document.querySelectorAll('.level-tab-btn').forEach(btn => {{
                btn.classList.remove('active');
            }});
            document.getElementById('btn-tiktok-' + level).classList.add('active');

            // Hide/Show contents
            document.querySelectorAll('.platform-content').forEach(div => {{
                if (div.id.startsWith('tiktok-')) {{
                    div.classList.remove('active');
                }}
            }});
            document.getElementById('tiktok-' + level + '-content').classList.add('active');

            // Hide/Show sidebar menus
            document.querySelectorAll('.menu-container').forEach(menu => {{
                if (menu.id.startsWith('tiktok-')) {{
                    menu.style.display = 'none';
                }}
            }});
            document.getElementById('tiktok-' + level + '-menu').style.display = 'block';

            // Subtitle text update
            const levelText = level === 'basic' ? "Bản Cơ Bản" : "Bản Nâng Cao";
            document.getElementById('app-subtitle').innerText = "Cẩm nang Vận hành Bán hàng TikTok Shop Việt Nam v2026 (" + levelText + ")";

            // Highlight first chapter in newly opened menu
            const activeLink = document.querySelector('#tiktok-' + level + '-menu .menu-link');
            document.querySelectorAll('.menu-link').forEach(link => link.classList.remove('active'));
            if (activeLink) activeLink.classList.add('active');

            // Scroll content area back to top
            document.getElementById('content-area').scrollTop = 0;

            // Clear search field
            document.getElementById('search-input').value = '';
            searchContent();
        }}

        // ----------------------------------------------------
        // 5. Theme Toggling
        // ----------------------------------------------------
        function toggleTheme() {{
            const html = document.documentElement;
            const currentTheme = html.getAttribute('data-theme');
            const newTheme = currentTheme === 'dark' ? 'light' : 'dark';
            html.setAttribute('data-theme', newTheme);
            
            const icon = document.getElementById('theme-icon');
            const text = document.getElementById('theme-text');
            
            if (newTheme === 'light') {{
                icon.innerHTML = `<path d="M12 7a5 5 0 1 1 0 10 5 5 0 0 1 0-10zm0 2a3 3 0 1 0 0 6 3 3 0 0 0 0-6zm0-7a1 1 0 0 1 1 1v1a1 1 0 1 1-2 0V3a1 1 0 0 1 1-1zm0 17a1 1 0 0 1 1 1v1a1 1 0 1 1-2 0v-1a1 1 0 0 1 1-1zm-8-8a1 1 0 0 1 1-1h1a1 1 0 1 1 0 2H5a1 1 0 0 1-1-1zm14 0a1 1 0 0 1 1-1h1a1 1 0 1 1 0 2h-1a1 1 0 0 1-1-1zM6.34 5.64a1 1 0 0 1 1.42 0l.7.7a1 1 0 1 1-1.42 1.42l-.7-.7a1 1 0 0 1 0-1.42zm10.6 10.6a1 1 0 0 1 1.42 0l.7.7a1 1 0 1 1-1.42 1.42l-.7-.7a1 1 0 0 1 0-1.42zM5.64 17.66a1 1 0 0 1 0-1.42l.7-.7a1 1 0 1 1 1.42 1.42l-.7.7a1 1 0 0 1-1.42 0zm10.6-10.6a1 1 0 0 1 0-1.42l.7-.7a1 1 0 1 1 1.42 1.42l-.7.7a1 1 0 0 1-1.42 0z"/>`;
                text.innerText = "Chế độ Tối";
            }} else {{
                icon.innerHTML = `<path d="M12 3a9 9 0 1 0 9 9c0-.46-.04-.92-.1-1.36a5.389 5.389 0 0 1-4.4 2.26 5.403 5.403 0 0 1-3.14-9.8c.44-.06.9-.1 1.36-.1H12z"/>`;
                text.innerText = "Chế độ Sáng";
            }}
        }}

        // ----------------------------------------------------
        // 6. Navigation Scrolling & Accordion Submenus
        // ----------------------------------------------------
        function toggleSidebarSubmenu(id) {{
            const submenu = document.getElementById('submenu-' + id);
            const parentLi = document.getElementById('menu-item-' + id);
            if (submenu && parentLi) {{
                const arrow = parentLi.querySelector('.submenu-arrow');
                if (submenu.style.display === 'none') {{
                    submenu.style.display = 'block';
                    if (arrow) arrow.innerText = '▼';
                }} else {{
                    submenu.style.display = 'none';
                    if (arrow) arrow.innerText = '▶';
                }}
            }}
        }}

        function expandShopeeSubmenu(activeChNum) {{
            for (let i = 1; i <= 6; i++) {{
                const submenu = document.getElementById('submenu-shopee-ch' + i);
                const parentLi = document.getElementById('menu-item-shopee-ch' + i);
                if (submenu && parentLi) {{
                    const arrow = parentLi.querySelector('.submenu-arrow');
                    if (i.toString() === activeChNum.toString()) {{
                        submenu.style.display = 'block';
                        if (arrow) arrow.innerText = '▼';
                    }} else {{
                        submenu.style.display = 'none';
                        if (arrow) arrow.innerText = '▶';
                    }}
                }}
            }}
        }}

        function scrollToSection(id) {{
            const element = document.getElementById(id);
            if (element) {{
                element.scrollIntoView({{ behavior: 'auto' }});
                
                const platform = document.documentElement.getAttribute('data-platform');
                if (platform === 'shopee') {{
                    // Update chapter highlight
                    document.querySelectorAll('#shopee-menu .menu-link').forEach(link => {{
                        link.classList.remove('active');
                    }});
                    // Update subsection highlight
                    document.querySelectorAll('#shopee-menu .submenu-link').forEach(link => {{
                        link.classList.remove('active');
                    }});
                    
                    if (id.startsWith('shopee-ch')) {{
                        const subLink = document.getElementById('sublink-' + id);
                        if (subLink) {{
                            subLink.classList.add('active');
                            // Highlight the parent menu-link
                            const parentLi = subLink.closest('.has-submenu');
                            if (parentLi) {{
                                const mainLink = parentLi.querySelector('.menu-link');
                                if (mainLink) mainLink.classList.add('active');
                            }}
                        }}
                    }} else if (id.startsWith('h1-chuong')) {{
                        const chNum = id.replace('h1-chuong', '');
                        const parentLi = document.getElementById('menu-item-shopee-ch' + chNum);
                        if (parentLi) {{
                            const mainLink = parentLi.querySelector('.menu-link');
                            if (mainLink) mainLink.classList.add('active');
                        }}
                    }}
                }} else {{
                    const menuId = 'tiktok-' + activeTikTokLevel + '-menu';
                    document.querySelectorAll('#' + menuId + ' .menu-link').forEach(link => {{
                        link.classList.remove('active');
                        const wrapper = link.closest('.menu-header-wrapper');
                        if (wrapper && wrapper.getAttribute('onclick') && wrapper.getAttribute('onclick').includes(id)) {{
                            link.classList.add('active');
                        }}
                    }});
                }}
            }}
        }}

        // ----------------------------------------------------
        // 7. Scroll Spy for active section sync
        // ----------------------------------------------------
        const contentArea = document.getElementById('content-area');
        contentArea.addEventListener('scroll', () => {{
            const platform = document.documentElement.getAttribute('data-platform');
            if (platform === 'shopee') {{
                // 1. Find active Chapter
                const chapters = document.querySelectorAll('.shopee-section');
                let activeChId = '';
                chapters.forEach(ch => {{
                    const rect = ch.getBoundingClientRect();
                    if (rect.top <= 250) {{
                        activeChId = ch.getAttribute('id');
                    }}
                }});

                // Update active state of chapter links in sidebar
                document.querySelectorAll('#shopee-menu .menu-link').forEach(link => {{
                    link.classList.remove('active');
                }});
                
                if (activeChId) {{
                    const chNum = activeChId.replace('h1-chuong', '');
                    const parentLi = document.getElementById('menu-item-shopee-ch' + chNum);
                    if (parentLi) {{
                        const mainLink = parentLi.querySelector('.menu-link');
                        if (mainLink) mainLink.classList.add('active');
                        expandShopeeSubmenu(chNum);
                    }}
                }}

                // 2. Find active Subsection
                const subsections = document.querySelectorAll('.shopee-sec-header');
                let activeSecId = '';
                subsections.forEach(sec => {{
                    const rect = sec.getBoundingClientRect();
                    if (rect.top <= 250) {{
                        activeSecId = sec.getAttribute('id');
                    }}
                }});

                document.querySelectorAll('#shopee-menu .submenu-link').forEach(link => {{
                    link.classList.remove('active');
                }});
                if (activeSecId) {{
                    const subLink = document.getElementById('sublink-' + activeSecId);
                    if (subLink) subLink.classList.add('active');
                }}
            }} else {{
                const h1s = document.querySelectorAll('.tiktok-section-' + activeTikTokLevel);
                const menuId = 'tiktok-' + activeTikTokLevel + '-menu';
                
                let current = '';
                h1s.forEach(h1 => {{
                    const rect = h1.getBoundingClientRect();
                    if (rect.top <= 200) {{
                        current = h1.getAttribute('id');
                    }}
                }});
                
                if (current) {{
                    document.querySelectorAll('#' + menuId + ' .menu-link').forEach(link => {{
                        link.classList.remove('active');
                        const wrapper = link.closest('.menu-header-wrapper');
                        if (wrapper && wrapper.getAttribute('onclick') && wrapper.getAttribute('onclick').includes(current)) {{
                            link.classList.add('active');
                        }}
                    }});
                }}
            }}
        }});

        // ----------------------------------------------------
        // 8. Search Engine with diacritics removal
        // ----------------------------------------------------
        function removeVietnameseDiacritics(str) {{
            return str
                .normalize('NFD')
                .replace(/[\u0300-\u036f]/g, '')
                .replace(/đ/g, 'd')
                .replace(/Đ/g, 'D')
                .toLowerCase();
        }}

        function searchContent() {{
            const rawQuery = document.getElementById('search-input').value;
            const queryNormalized = removeVietnameseDiacritics(rawQuery).trim();
            const platform = document.documentElement.getAttribute('data-platform');
            
            let containerId = platform === 'shopee' ? 'shopee-content' : ('tiktok-' + activeTikTokLevel + '-content');
            const body = document.getElementById(containerId);
            
            const paragraphs = body.querySelectorAll('p, li, h1, h2, h3, h4, table tr, .widget-card');
            
            paragraphs.forEach(p => {{
                const text = p.innerText || p.textContent;
                const textNormalized = removeVietnameseDiacritics(text);
                
                if (textNormalized.includes(queryNormalized)) {{
                    p.style.display = '';
                    if (queryNormalized !== '') {{
                        p.style.backgroundColor = 'rgba(238, 77, 45, 0.1)';
                    }} else {{
                        p.style.backgroundColor = '';
                    }}
                }} else {{
                    if (p.classList.contains('widget-card') || p.tagName === 'TR') {{
                        p.style.display = '';
                    }} else {{
                        p.style.display = 'none';
                    }}
                }}
            }});
        }}

        // ----------------------------------------------------
        // 9. Lightbox zoomable images
        // ----------------------------------------------------
        const modal = document.getElementById('lightbox');
        const modalImg = document.getElementById('lightbox-img');
        const captionText = document.getElementById('lightbox-caption');
        
        function setupLightboxListeners() {{
            document.querySelectorAll('.content-body img').forEach(img => {{
                img.onclick = function() {{
                    modal.classList.add('show');
                    modalImg.src = this.src;
                    modalImg.style.maxHeight = '80vh';
                    captionText.innerHTML = this.alt || "Hướng dẫn minh họa vận hành";
                }}
            }});
        }}
        setupLightboxListeners();

        function closeLightbox() {{
            modal.classList.remove('show');
        }}

        // ----------------------------------------------------
        // 10. Shopee Onboarding Checklist Widget
        // ----------------------------------------------------
        const shopeeTotalItems = 6;
        let shopeeChecklistState = [false, false, false, false, false, false];

        if (localStorage.getItem('shopee_onboarding_checklist')) {{
            try {{
                shopeeChecklistState = JSON.parse(localStorage.getItem('shopee_onboarding_checklist'));
            }} catch (e) {{
                console.error(e);
            }}
        }}

        function initShopeeChecklist() {{
            for (let i = 0; i < shopeeTotalItems; i++) {{
                const chk = document.getElementById('chk-' + i);
                if (chk) {{
                    chk.checked = shopeeChecklistState[i];
                    const label = chk.parentElement;
                    if (shopeeChecklistState[i]) {{
                        label.classList.add('checked');
                    }} else {{
                        label.classList.remove('checked');
                    }}
                }}
            }}
            updateShopeeChecklistProgress();
        }}

        function toggleChecklistItem(index) {{
            const chk = document.getElementById('chk-' + index);
            shopeeChecklistState[index] = chk.checked;
            const label = chk.parentElement;
            if (chk.checked) {{
                label.classList.add('checked');
            }} else {{
                label.classList.remove('checked');
            }}
            localStorage.setItem('shopee_onboarding_checklist', JSON.stringify(shopeeChecklistState));
            updateShopeeChecklistProgress();
        }}

        function updateShopeeChecklistProgress() {{
            const checkedCount = shopeeChecklistState.filter(Boolean).length;
            const pct = Math.round((checkedCount / shopeeTotalItems) * 100);
            
            const pctEl = document.getElementById('checklist-pct');
            if (pctEl) pctEl.innerText = "Tiến độ: " + pct + "% (" + checkedCount + "/" + shopeeTotalItems + " bước)";
            
            const fillEl = document.getElementById('checklist-progress-fill');
            if (fillEl) fillEl.style.width = pct + "%";
        }}

        initShopeeChecklist();

        // ----------------------------------------------------
        // 11. TikTok Shop Onboarding Checklist Widget
        // ----------------------------------------------------
        const tiktokTotalItems = 6;
        let tiktokChecklistStateBasic = [false, false, false, false, false, false];
        let tiktokChecklistStateAdvanced = [false, false, false, false, false, false];

        // Load basic checklist
        if (localStorage.getItem('tiktok_onboarding_checklist_basic')) {{
            try {{
                tiktokChecklistStateBasic = JSON.parse(localStorage.getItem('tiktok_onboarding_checklist_basic'));
            }} catch (e) {{}}
        }}
        // Load advanced checklist
        if (localStorage.getItem('tiktok_onboarding_checklist_advanced')) {{
            try {{
                tiktokChecklistStateAdvanced = JSON.parse(localStorage.getItem('tiktok_onboarding_checklist_advanced'));
            }} catch (e) {{}}
        }}

        function initTikTokChecklist(level) {{
            const state = level === 'basic' ? tiktokChecklistStateBasic : tiktokChecklistStateAdvanced;
            for (let i = 0; i < tiktokTotalItems; i++) {{
                const chk = document.getElementById('tt-chk-' + level + '-' + i);
                if (chk) {{
                    chk.checked = state[i];
                    const label = chk.parentElement;
                    if (state[i]) {{
                        label.classList.add('checked');
                    }} else {{
                        label.classList.remove('checked');
                    }}
                }}
            }}
            updateTikTokChecklistProgress(level);
        }}

        function toggleTikTokChecklistItem(level, index) {{
            const state = level === 'basic' ? tiktokChecklistStateBasic : tiktokChecklistStateAdvanced;
            const chk = document.getElementById('tt-chk-' + level + '-' + index);
            state[index] = chk.checked;
            const label = chk.parentElement;
            if (chk.checked) {{
                label.classList.add('checked');
            }} else {{
                label.classList.remove('checked');
            }}
            localStorage.setItem('tiktok_onboarding_checklist_' + level, JSON.stringify(state));
            updateTikTokChecklistProgress(level);
        }}

        function updateTikTokChecklistProgress(level) {{
            const state = level === 'basic' ? tiktokChecklistStateBasic : tiktokChecklistStateAdvanced;
            const checkedCount = state.filter(Boolean).length;
            const pct = Math.round((checkedCount / tiktokTotalItems) * 100);
            
            const pctEl = document.getElementById('tiktok-checklist-pct-' + level);
            if (pctEl) pctEl.innerText = "Tiến độ: " + pct + "% (" + checkedCount + "/" + tiktokTotalItems + " bước)";
            
            const fillEl = document.getElementById('tiktok-checklist-progress-fill-' + level);
            if (fillEl) fillEl.style.width = pct + "%";
        }}

        initTikTokChecklist('basic');
        initTikTokChecklist('advanced');

        // ----------------------------------------------------
        // 12. OBS Livestream Config Profile
        // ----------------------------------------------------
        const obsProfiles = {{
            low: {{
                res: "540p (960 x 540) - Thích hợp kết nối mạng gia đình trung bình",
                fps: "30 FPS",
                bitrate: "1000 - 1500 Kbps",
                encoder: "x264 (Phần mềm CPU) hoặc QSV (Intel)"
            }},
            mid: {{
                res: "720p (1280 x 720) - Khuyến nghị cho luồng Live rõ nét",
                fps: "30 FPS hoặc 60 FPS",
                bitrate: "2000 - 2500 Kbps",
                encoder: "NVIDIA NVENC H.264 (Card đồ họa rời Geforce) hoặc AMD HW H.264"
            }},
            high: {{
                res: "1080p (1920 x 1080) - Chất lượng cao sắc nét (Cần cáp quang ổn định)",
                fps: "60 FPS",
                bitrate: "3000 - 4500 Kbps",
                encoder: "NVIDIA NVENC H.264 (New) hoặc Card đồ họa rời chuyên dụng"
            }}
        }};

        function selectOBSProfile(profileName) {{
            const btns = document.querySelectorAll('.profile-btn');
            btns.forEach(btn => btn.classList.remove('active'));
            event.target.classList.add('active');
            
            const config = obsProfiles[profileName];
            document.getElementById('obs-res').innerText = config.res;
            document.getElementById('obs-fps').innerText = config.fps;
            document.getElementById('obs-bitrate').innerText = config.bitrate;
            document.getElementById('obs-encoder').innerText = config.encoder;
        }}

        function copyOBSSettings() {{
            const res = document.getElementById('obs-res').innerText;
            const fps = document.getElementById('obs-fps').innerText;
            const bitrate = document.getElementById('obs-bitrate').innerText;
            const encoder = document.getElementById('obs-encoder').innerText;
            
            const settingsText = "Thông số cấu hình OBS Livestream Shopee khuyến nghị:\\n" +
                                 "- Độ phân giải: " + res + "\\n" +
                                 "- Tỷ lệ khung hình: " + fps + "\\n" +
                                 "- Bitrate Video: " + bitrate + "\\n" +
                                 "- Bộ mã hóa (Encoder): " + encoder + "\\n" +
                                 "- Keyframe Interval: 2 giây (Bắt buộc)";
                                 
            navigator.clipboard.writeText(settingsText).then(() => {{
                alert("Đã sao chép cấu hình OBS vào khay nhớ tạm!");
            }});
        }}

        // ----------------------------------------------------
        // 13. Return/Refund Dispute Assistant
        // ----------------------------------------------------
        const disputeDatabase = {{
            empty_box: {{
                evidence: [
                    "<strong>Video đóng gói sản phẩm của Shop:</strong> Quay rõ mã vận đơn đơn hàng, tình trạng sản phẩm trước khi bọc, quy trình dán băng keo niêm phong.",
                    "<strong>Biên bản bàn giao cho bưu tá / Phiếu gửi bưu cục:</strong> Có chữ ký biên nhận của nhân viên giao nhận.",
                    "<strong>Hình ảnh cân nặng sản phẩm khi đóng gói:</strong> (Nếu có chụp cân nặng trên cân điện tử)."
                ],
                template: "Kính gửi Shopee, Tôi là Người bán của đơn hàng [MÃ_ĐƠN_HÀNG]. Khách hàng khiếu nại nhận được hộp rỗng. Tuy nhiên, shop tôi có cung cấp video đóng gói sản phẩm ghi nhận quá trình đóng gói sản phẩm nguyên vẹn, dán tem mã vận đơn và giao cho ĐVVC đúng quy cách. Trọng lượng kiện hàng khi gửi đi là [SỐ_KG] kg. Kính mong Shopee đối soát cân nặng chặng đi của ĐVVC để làm rõ thất thoát và hoàn lại số tiền đơn hàng cho shop. Xin cảm ơn."
            }},
            missing: {{
                evidence: [
                    "<strong>Video đóng gói sản phẩm:</strong> Quay rõ danh sách các sản phẩm đặt vào hộp tương thích với hóa đơn/mã đơn hàng.",
                    "<strong>Video khui hàng hoàn trả (nếu có):</strong> Để chứng minh shop nhận lại những sản phẩm nào từ người mua.",
                    "<strong>Hình ảnh cân nặng kiện hàng trước khi gửi đi.</strong>"
                ],
                template: "Kính gửi Shopee, Tôi xin khiếu nại yêu cầu Trả hàng/Hoàn tiền đơn hàng [MÃ_ĐƠN_HÀNG] với lý do thiếu hàng. Theo video đóng gói đính kèm, shop tôi đã đặt đầy đủ [TÊN_SẢN_PHẨM_THIẾU] cùng với quà tặng đi kèm vào trong kiện hàng trước khi niêm phong. Cân nặng đóng gói của đơn hàng hiển thị trên phiếu gửi là [SỐ_KG] kg. Kính đề nghị Shopee kiểm tra và từ chối yêu cầu hoàn tiền của Người mua. Xin cảm ơn."
            }},
            damaged: {{
                evidence: [
                    "<strong>Video đóng gói sản phẩm:</strong> Chứng minh sản phẩm nguyên vẹn, không bể vỡ và bọc chống sốc dầy dặn từ 3-4 lớp.",
                    "<strong>Hình ảnh 6 mặt hộp ngoài của khách gửi hoàn trả:</strong> Cho thấy kiện hàng bị móp méo, rách nát do khâu vận chuyển đè ép.",
                    "<strong>Video khui hàng hoàn:</strong> Ghi rõ trạng thái sản phẩm bên trong bị dập nát, gãy hỏng."
                ],
                template: "Kính gửi Shopee, Shop tôi muốn khiếu nại về tình trạng hư hỏng của đơn hàng [MÃ_ĐƠN_HÀNG]. Khi gửi đi, sản phẩm hoàn toàn nguyên vẹn và được bọc chống sốc bảo vệ kỹ lượng (có video đóng gói đính kèm). Sản phẩm bị vỡ dập là do lỗi vận chuyển của ĐVVC đè nén sản phẩm trong chặng giao. Kính mong Shopee hỗ trợ đền bù giá trị đơn hàng cho shop theo chính sách bảo hiểm mất mát/hư hỏng của sàn. Xin cảm ơn."
            }},
            wrong_item: {{
                evidence: [
                    "<strong>Video đóng gói của shop:</strong> Chứng minh sản phẩm gửi đi đúng mẫu mã, màu sắc, size khách đặt.",
                    "<strong>Video khui kiện hàng hoàn trả:</strong> Cho thấy sản phẩm khách gửi trả về là một sản phẩm khác hoàn toàn (khác thương hiệu, sản phẩm cũ nát, hoặc gạch đá, giấy lộn...).",
                    "<strong>Hình ảnh mã vận đơn hoàn trả dán đè lên hộp cũ.</strong>"
                ],
                template: "Kính gửi Shopee, Shop khiếu nại đơn hàng hoàn trả [MÃ_ĐƠN_HÀNG] bị tráo hàng / sai mẫu. Shop đã gửi đúng [TÊN_SẢN_PHẨM_GỐC] (có video đóng gói đính kèm). Tuy nhiên, sản phẩm hoàn trả nhận được từ người mua lại là [TÊN_SẢN_PHẨM_KHÁCH_TRẢ_VỀ] (cũ nát/không phải sản phẩm của shop). Video khui hàng hoàn cho thấy rõ mã vận đơn hoàn trả và quá trình mở hộp. Đề nghị Shopee từ chối hoàn tiền cho Người mua và bồi thường đơn hàng cho shop. Xin cảm ơn."
            }}
        }};

        function generateDisputeTemplate() {{
            const reason = document.getElementById('dispute-reason-select').value;
            const data = disputeDatabase[reason];
            const listEl = document.getElementById('dispute-evidence-list');
            listEl.innerHTML = '';
            data.evidence.forEach(ev => {{
                const li = document.createElement('li');
                li.innerHTML = ev;
                listEl.appendChild(li);
            }});
            document.getElementById('dispute-text-template').value = data.template;
        }}

        generateDisputeTemplate();

        // ----------------------------------------------------
        // 14. Shopee Tax Calculator (NĐ 117)
        // ----------------------------------------------------
        function calculateTax() {{
            const price = parseFloat(document.getElementById('calc-price').value) || 0;
            const voucherSeller = parseFloat(document.getElementById('calc-voucher-seller').value) || 0;
            const voucherCo = parseFloat(document.getElementById('calc-voucher-co').value) || 0;
            const category = document.getElementById('calc-category').value;
            
            const taxableRevenue = Math.max(0, price - voucherSeller - voucherCo);
            let vatRate = 0.01;
            let pitRate = 0.005;
            
            if (category !== 'goods') {{
                vatRate = 0.05;
                pitRate = 0.02;
            }}
            
            const vat = taxableRevenue * vatRate;
            const pit = taxableRevenue * pitRate;
            const totalTax = vat + pit;
            
            document.getElementById('res-taxable-revenue').innerText = Math.round(taxableRevenue).toLocaleString('vi-VN') + " VNĐ";
            document.getElementById('res-vat').innerText = Math.round(vat).toLocaleString('vi-VN') + " VNĐ";
            document.getElementById('res-pit').innerText = Math.round(pit).toLocaleString('vi-VN') + " VNĐ";
            document.getElementById('res-total-tax').innerText = Math.round(totalTax).toLocaleString('vi-VN') + " VNĐ";
        }}
        calculateTax();

        // ----------------------------------------------------
        // 15. TikTok Shop Policy Word Checker Widget
        // ----------------------------------------------------
        const forbiddenWords = {{
            "chữa": "hỗ trợ / bổ sung",
            "trị": "cải thiện / chăm sóc",
            "khỏi": "phục hồi / cải thiện tốt hơn",
            "đặc trị": "chuyên biệt / hỗ trợ giảm",
            "hết bệnh": "cơ thể khỏe mạnh hơn",
            "cam kết": "đảm bảo / đồng hành hướng dẫn",
            "100%": "phần lớn / tối ưu theo cơ địa",
            "chắc chắn": "hỗ trợ rất tốt",
            "tốt nhất": "được đánh giá cao / uy tín",
            "số 1": "phổ biến / thuộc nhóm đầu",
            "thần dược": "thảo dược / thực phẩm hỗ trợ",
            "cứu tinh": "lựa chọn tốt",
            "bác sĩ khuyên dùng": "được tham khảo ý kiến y tế",
            "bệnh viện tin dùng": "đáp ứng tiêu chuẩn kiểm định"
        }};

        function checkTikTokPolicy() {{
            const text = document.getElementById('policy-checker-input').value.toLowerCase();
            const resBox = document.getElementById('policy-checker-result-box');
            const listEl = document.getElementById('policy-checker-violations');
            
            let detected = [];
            
            for (let word in forbiddenWords) {{
                if (text.includes(word)) {{
                    detected.push({{ word: word, replacement: forbiddenWords[word] }});
                }}
            }}
            
            if (detected.length > 0) {{
                resBox.style.display = 'block';
                listEl.innerHTML = '';
                detected.forEach(item => {{
                    const div = document.createElement('div');
                    div.style.marginBottom = '6px';
                    div.innerHTML = `<span class="policy-tag danger">${{item.word}}</span> -> thay bằng: <span class="policy-tag safe">${{item.replacement}}</span>`;
                    listEl.appendChild(div);
                }});
            }} else {{
                resBox.style.display = 'none';
            }}
        }}

        function cleanTextAndReplace() {{
            let text = document.getElementById('policy-checker-input').value;
            let cleanText = text;
            
            for (let word in forbiddenWords) {{
                const regex = new RegExp(word, 'gi');
                cleanText = cleanText.replace(regex, forbiddenWords[word]);
            }}
            
            document.getElementById('policy-checker-input').value = cleanText;
            checkTikTokPolicy();
            alert("Đã tự động sửa các từ rủi ro sang từ an toàn!");
        }}

        // ----------------------------------------------------
        // 16. Multi-row Campaign profit calculators
        // ----------------------------------------------------
        
        // --- SHOPEE CALC GRID ---
        function renderShopeeTable() {{
            const tbody = document.getElementById('shopee-calc-tbody');
            if (!tbody) return;
            tbody.innerHTML = '';
            
            shopeeCalcRows.forEach((row, index) => {{
                const prod = shopeeProducts.find(p => p.sku === row.sku) || {{
                    sku: row.sku,
                    name: "Sản phẩm tùy chỉnh",
                    comm_regular: 0.1,
                    comm_mall: 0.12,
                    ref_price: 0,
                    cogs: 0,
                    packing: 0,
                    default_cost: 0
                }};
                
                if (row.cogs === null || row.cogs === undefined || isNaN(row.cogs)) {{
                    if (prod.cogs !== null && prod.cogs !== undefined && prod.cogs > 0) {{
                        row.cogs = prod.cogs + (prod.packing || 0);
                    }} else if (prod.default_cost > 0) {{
                        row.cogs = prod.default_cost;
                    }} else {{
                        row.cogs = (prod.ref_price * 0.7) + (prod.packing || 0);
                    }}
                }}
                
                const price = row.price;
                const voucher = row.voucher;
                const affiliateRate = row.affiliateRate;
                const shopType = row.shopType;
                const cogs = row.cogs;
                const other = row.other || 0;
                
                const rev = Math.max(0, price - voucher);
                const commRate = shopType === 'mall' ? prod.comm_mall : prod.comm_regular;
                const commFee = rev * commRate;
                const transFee = rev * 0.06; // Fixed 6%
                const handlingFee = 3000;
                const totalFees = commFee + transFee + handlingFee;
                
                const affiliateFee = rev * (affiliateRate / 100);
                const totalCost = cogs + affiliateFee + other + totalFees;
                const profit = rev - totalCost;
                const margin = rev > 0 ? (profit / rev) * 100 : 0;
                
                let badgeColor = '#ef4444';
                let badgeText = 'LỖ';
                if (profit >= 0) {{
                    if (margin < 10) {{
                        badgeColor = '#f59e0b';
                        badgeText = 'LÃI THẤP';
                    }} else if (margin < 20) {{
                        badgeColor = '#3b82f6';
                        badgeText = 'ỔN';
                    }} else {{
                        badgeColor = '#22c55e';
                        badgeText = 'TỐT';
                    }}
                }}
                
                let optionsHtml = '';
                shopeeProducts.forEach(p => {{
                    const selected = p.sku === row.sku ? 'selected' : '';
                    optionsHtml += `<option value="${{p.sku}}" ${{selected}}>[${{p.sku}}] ${{p.name.substring(0, 30)}}...</option>`;
                }});
                
                const tr = document.createElement('tr');
                tr.innerHTML = `
                    <td style="padding: 6px 10px;">
                        <select onchange="updateShopeeRowProduct(${{index}}, this.value)" style="width:100%; padding:6px; font-size:12px; background:var(--bg-primary); border:1px solid var(--border-color); border-radius:4px; color:var(--text-primary); outline:none;">
                            ${{optionsHtml}}
                        </select>
                        <div style="font-size: 10px; color: var(--text-secondary); margin-top: 4px; max-width: 250px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap;">
                            ${{prod.name}}
                        </div>
                    </td>
                    <td style="padding: 6px 10px;">
                        <select onchange="updateShopeeRowField(${{index}}, 'shopType', this.value)" style="width:100%; padding:6px; font-size:12px; background:var(--bg-primary); border:1px solid var(--border-color); border-radius:4px; color:var(--text-primary); outline:none;">
                            <option value="regular" ${{shopType === 'regular' ? 'selected' : ''}}>Thường</option>
                            <option value="mall" ${{shopType === 'mall' ? 'selected' : ''}}>Mall</option>
                        </select>
                    </td>
                    <td style="padding: 6px 10px;">
                        <input type="number" value="${{price}}" oninput="updateShopeeRowField(${{index}}, 'price', parseFloat(this.value) || 0)" style="width:100%; padding:6px; font-size:12px; background:var(--bg-primary); border:1px solid var(--border-color); border-radius:4px; color:var(--text-primary); outline:none;">
                    </td>
                    <td style="padding: 6px 10px;">
                        <input type="number" value="${{voucher}}" oninput="updateShopeeRowField(${{index}}, 'voucher', parseFloat(this.value) || 0)" style="width:100%; padding:6px; font-size:12px; background:var(--bg-primary); border:1px solid var(--border-color); border-radius:4px; color:var(--text-primary); outline:none;">
                    </td>
                    <td style="padding: 6px 10px;">
                        <input type="number" value="${{affiliateRate}}" oninput="updateShopeeRowField(${{index}}, 'affiliateRate', parseFloat(this.value) || 0)" style="width:100%; padding:6px; font-size:12px; background:var(--bg-primary); border:1px solid var(--border-color); border-radius:4px; color:var(--text-primary); outline:none;">
                    </td>
                    <td style="padding: 6px 10px;">
                        <input type="number" value="${{cogs}}" oninput="updateShopeeRowField(${{index}}, 'cogs', parseFloat(this.value) || 0)" style="width:100%; padding:6px; font-size:12px; background:var(--bg-primary); border:1px solid var(--border-color); border-radius:4px; color:var(--text-primary); outline:none;">
                    </td>
                    <td style="padding: 6px 10px;">
                        <input type="number" value="${{other}}" oninput="updateShopeeRowField(${{index}}, 'other', parseFloat(this.value) || 0)" placeholder="Ads, ship..." style="width:100%; padding:6px; font-size:12px; background:var(--bg-primary); border:1px solid var(--border-color); border-radius:4px; color:var(--text-primary); outline:none;">
                    </td>
                    <td style="padding: 6px 10px; font-size:12px; vertical-align: middle;" title="Hoa hồng (${{(commRate*100).toFixed(1)}}%): ${{Math.round(commFee).toLocaleString()}}đ&#10;Giao dịch (6.0%): ${{Math.round(transFee).toLocaleString()}}đ&#10;Xử lý cố định: 3,000đ">
                        <div style="font-weight:600; color:#ef4444; cursor:help;">${{Math.round(totalFees).toLocaleString()}}đ</div>
                        <div style="font-size:9px; color:var(--text-secondary);">Hover xem chi tiết</div>
                    </td>
                    <td style="padding: 6px 10px; font-size:12px; vertical-align: middle;">
                        <div style="font-weight:700; color:${{profit >= 0 ? '#22c55e' : '#ef4444'}};">${{Math.round(profit).toLocaleString()}}đ</div>
                        <span style="font-size:9.5px; background:${{badgeColor}}20; color:${{badgeColor}}; border:1px solid ${{badgeColor}}; padding:1px 3px; border-radius:3px; font-weight:600; display:inline-block; margin-top:2px;">${{badgeText}} (${{margin.toFixed(1)}}%)</span>
                    </td>
                    <td style="padding: 6px 10px; text-align:center; vertical-align: middle;">
                        <button onclick="deleteShopeeRow(${{index}})" style="background:none; border:none; color:var(--text-secondary); cursor:pointer; font-size:16px; font-weight:bold;">&times;</button>
                    </td>
                `;
                tbody.appendChild(tr);
            }});
            
            updateShopeeDashboard();
        }}

        function updateShopeeRowProduct(index, sku) {{
            const prod = shopeeProducts.find(p => p.sku === sku);
            shopeeCalcRows[index].sku = sku;
            shopeeCalcRows[index].price = prod.ref_price || 200000;
            shopeeCalcRows[index].cogs = null; // Forces recalculation
            localStorage.setItem('shopee_calc_rows', JSON.stringify(shopeeCalcRows));
            renderShopeeTable();
        }}

        function updateShopeeRowField(index, field, val) {{
            shopeeCalcRows[index][field] = val;
            localStorage.setItem('shopee_calc_rows', JSON.stringify(shopeeCalcRows));
            updateShopeeDashboard();
            renderShopeeTable();
        }}

        function addShopeeCalcRow() {{
            if (shopeeProducts.length === 0) return;
            shopeeCalcRows.push({{
                sku: shopeeProducts[0].sku,
                shopType: 'regular',
                price: shopeeProducts[0].ref_price || 200000,
                voucher: 0,
                affiliateRate: 0,
                cogs: null,
                other: 0
            }});
            localStorage.setItem('shopee_calc_rows', JSON.stringify(shopeeCalcRows));
            renderShopeeTable();
        }}

        function deleteShopeeRow(index) {{
            shopeeCalcRows.splice(index, 1);
            localStorage.setItem('shopee_calc_rows', JSON.stringify(shopeeCalcRows));
            renderShopeeTable();
        }}

        function updateShopeeDashboard() {{
            let totalRevenue = 0;
            let totalFees = 0;
            let totalCogsAndOther = 0;
            let totalProfit = 0;
            
            shopeeCalcRows.forEach(row => {{
                const prod = shopeeProducts.find(p => p.sku === row.sku) || {{}};
                const rev = Math.max(0, row.price - row.voucher);
                const commRate = row.shopType === 'mall' ? prod.comm_mall : prod.comm_regular;
                const fees = (rev * commRate) + (rev * 0.06) + 3000;
                
                const cogs = row.cogs || 0;
                const affiliateFee = rev * (row.affiliateRate / 100);
                const other = row.other || 0;
                
                totalRevenue += rev;
                totalFees += fees;
                totalCogsAndOther += cogs + affiliateFee + other;
                totalProfit += (rev - (cogs + affiliateFee + other + fees));
            }});
            
            const margin = totalRevenue > 0 ? (totalProfit / totalRevenue) * 100 : 0;
            
            document.getElementById('sp-dash-revenue').innerText = Math.round(totalRevenue).toLocaleString('vi-VN') + " VNĐ";
            document.getElementById('sp-dash-fees').innerText = Math.round(totalFees).toLocaleString('vi-VN') + " VNĐ";
            document.getElementById('sp-dash-cogs').innerText = Math.round(totalCogsAndOther).toLocaleString('vi-VN') + " VNĐ";
            
            const profitEl = document.getElementById('sp-dash-profit');
            profitEl.innerText = Math.round(totalProfit).toLocaleString('vi-VN') + " VNĐ";
            profitEl.style.color = totalProfit >= 0 ? '#22c55e' : '#ef4444';
            
            document.getElementById('sp-dash-margin').innerText = margin.toFixed(1) + "%";
        }}

        // --- TIKTOK CALC GRID ---
        function renderTikTokTable() {{
            const tbody = document.getElementById('tiktok-calc-tbody');
            if (!tbody) return;
            tbody.innerHTML = '';
            
            tiktokCalcRows.forEach((row, index) => {{
                const prod = tiktokProducts.find(p => p.sku === row.sku) || {{
                    sku: row.sku,
                    name: "Sản phẩm tùy chỉnh",
                    comm_regular: 0.1,
                    comm_mall: 0.12,
                    cogs: 0,
                    packing: 0,
                    gift: 0
                }};
                
                if (row.cogs === null || row.cogs === undefined || isNaN(row.cogs)) {{
                    row.cogs = prod.cogs + (prod.packing || 0) + (prod.gift || 0);
                }}
                
                const price = row.price;
                const voucher = row.voucher;
                const affiliateRate = row.affiliateRate;
                const shopType = row.shopType;
                const cogs = row.cogs;
                const other = row.other || 0;
                
                const rev = Math.max(0, price - voucher);
                const commRate = shopType === 'mall' ? prod.comm_mall : prod.comm_regular;
                const commFee = rev * commRate;
                const transFee = rev * 0.06; // Fixed 6%
                const handlingFee = 3000;
                const totalFees = commFee + transFee + handlingFee;
                
                const affiliateFee = rev * (affiliateRate / 100);
                const totalCost = cogs + affiliateFee + other + totalFees;
                const profit = rev - totalCost;
                const margin = rev > 0 ? (profit / rev) * 100 : 0;
                
                let badgeColor = '#ef4444';
                let badgeText = 'LỖ';
                if (profit >= 0) {{
                    if (margin < 10) {{
                        badgeColor = '#f59e0b';
                        badgeText = 'LÃI THẤP';
                    }} else if (margin < 15) {{
                        badgeColor = '#3b82f6';
                        badgeText = 'ỔN';
                    }} else {{
                        badgeColor = '#22c55e';
                        badgeText = 'TỐT';
                    }}
                }}
                
                let optionsHtml = '';
                tiktokProducts.forEach(p => {{
                    const selected = p.sku === row.sku ? 'selected' : '';
                    optionsHtml += `<option value="${{p.sku}}" ${{selected}}>[${{p.sku}}] ${{p.name.substring(0, 30)}}...</option>`;
                }});
                
                const tr = document.createElement('tr');
                tr.innerHTML = `
                    <td style="padding: 6px 10px;">
                        <select onchange="updateTikTokRowProduct(${{index}}, this.value)" style="width:100%; padding:6px; font-size:12px; background:var(--bg-primary); border:1px solid var(--border-color); border-radius:4px; color:var(--text-primary); outline:none;">
                            ${{optionsHtml}}
                        </select>
                        <div style="font-size: 10px; color: var(--text-secondary); margin-top: 4px; max-width: 250px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap;">
                            ${{prod.name}}
                        </div>
                    </td>
                    <td style="padding: 6px 10px;">
                        <select onchange="updateTikTokRowField(${{index}}, 'shopType', this.value)" style="width:100%; padding:6px; font-size:12px; background:var(--bg-primary); border:1px solid var(--border-color); border-radius:4px; color:var(--text-primary); outline:none;">
                            <option value="regular" ${{shopType === 'regular' ? 'selected' : ''}}>Thường</option>
                            <option value="mall" ${{shopType === 'mall' ? 'selected' : ''}}>Mall</option>
                        </select>
                    </td>
                    <td style="padding: 6px 10px;">
                        <input type="number" value="${{price}}" oninput="updateTikTokRowField(${{index}}, 'price', parseFloat(this.value) || 0)" style="width:100%; padding:6px; font-size:12px; background:var(--bg-primary); border:1px solid var(--border-color); border-radius:4px; color:var(--text-primary); outline:none;">
                    </td>
                    <td style="padding: 6px 10px;">
                        <input type="number" value="${{voucher}}" oninput="updateTikTokRowField(${{index}}, 'voucher', parseFloat(this.value) || 0)" style="width:100%; padding:6px; font-size:12px; background:var(--bg-primary); border:1px solid var(--border-color); border-radius:4px; color:var(--text-primary); outline:none;">
                    </td>
                    <td style="padding: 6px 10px;">
                        <input type="number" value="${{affiliateRate}}" oninput="updateTikTokRowField(${{index}}, 'affiliateRate', parseFloat(this.value) || 0)" style="width:100%; padding:6px; font-size:12px; background:var(--bg-primary); border:1px solid var(--border-color); border-radius:4px; color:var(--text-primary); outline:none;">
                    </td>
                    <td style="padding: 6px 10px;">
                        <input type="number" value="${{cogs}}" oninput="updateTikTokRowField(${{index}}, 'cogs', parseFloat(this.value) || 0)" style="width:100%; padding:6px; font-size:12px; background:var(--bg-primary); border:1px solid var(--border-color); border-radius:4px; color:var(--text-primary); outline:none;">
                    </td>
                    <td style="padding: 6px 10px;">
                        <input type="number" value="${{other}}" oninput="updateTikTokRowField(${{index}}, 'other', parseFloat(this.value) || 0)" placeholder="Ads, ship..." style="width:100%; padding:6px; font-size:12px; background:var(--bg-primary); border:1px solid var(--border-color); border-radius:4px; color:var(--text-primary); outline:none;">
                    </td>
                    <td style="padding: 6px 10px; font-size:12px; vertical-align: middle;" title="Hoa hồng (${{(commRate*100).toFixed(1)}}%): ${{Math.round(commFee).toLocaleString()}}đ&#10;Giao dịch (6.0%): ${{Math.round(transFee).toLocaleString()}}đ&#10;Xử lý cố định: 3,000đ">
                        <div style="font-weight:600; color:#ef4444; cursor:help;">${{Math.round(totalFees).toLocaleString()}}đ</div>
                        <div style="font-size:9px; color:var(--text-secondary);">Hover xem chi tiết</div>
                    </td>
                    <td style="padding: 6px 10px; font-size:12px; vertical-align: middle;">
                        <div style="font-weight:700; color:${{profit >= 0 ? '#22c55e' : '#ef4444'}};">${{Math.round(profit).toLocaleString()}}đ</div>
                        <span style="font-size:9.5px; background:${{badgeColor}}20; color:${{badgeColor}}; border:1px solid ${{badgeColor}}; padding:1px 3px; border-radius:3px; font-weight:600; display:inline-block; margin-top:2px;">${{badgeText}} (${{margin.toFixed(1)}}%)</span>
                    </td>
                    <td style="padding: 6px 10px; text-align:center; vertical-align: middle;">
                        <button onclick="deleteTikTokRow(${{index}})" style="background:none; border:none; color:var(--text-secondary); cursor:pointer; font-size:16px; font-weight:bold;">&times;</button>
                    </td>
                `;
                tbody.appendChild(tr);
            }});
            
            updateTikTokDashboard();
        }}

        function updateTikTokRowProduct(index, sku) {{
            const prod = tiktokProducts.find(p => p.sku === sku);
            tiktokCalcRows[index].sku = sku;
            tiktokCalcRows[index].cogs = null; // Forces recalculation
            localStorage.setItem('tiktok_calc_rows', JSON.stringify(tiktokCalcRows));
            renderTikTokTable();
        }}

        function updateTikTokRowField(index, field, val) {{
            tiktokCalcRows[index][field] = val;
            localStorage.setItem('tiktok_calc_rows', JSON.stringify(tiktokCalcRows));
            updateTikTokDashboard();
            renderTikTokTable();
        }}

        // Add default tiktok product helper
        function addTikTokCalcRow() {{
            if (tiktokProducts.length === 0) return;
            tiktokCalcRows.push({{
                sku: tiktokProducts[0].sku,
                shopType: 'regular',
                price: 200000,
                voucher: 0,
                affiliateRate: 0,
                cogs: null,
                other: 0
            }});
            localStorage.setItem('tiktok_calc_rows', JSON.stringify(tiktokCalcRows));
            renderTikTokTable();
        }}

        function deleteTikTokRow(index) {{
            tiktokCalcRows.splice(index, 1);
            localStorage.setItem('tiktok_calc_rows', JSON.stringify(tiktokCalcRows));
            renderTikTokTable();
        }}

        function updateTikTokDashboard() {{
            let totalRevenue = 0;
            let totalFees = 0;
            let totalCogsAndOther = 0;
            let totalProfit = 0;
            
            tiktokCalcRows.forEach(row => {{
                const prod = tiktokProducts.find(p => p.sku === row.sku) || {{}};
                const rev = Math.max(0, row.price - row.voucher);
                const commRate = row.shopType === 'mall' ? prod.comm_mall : prod.comm_regular;
                const fees = (rev * commRate) + (rev * 0.06) + 3000;
                
                const cogs = row.cogs || 0;
                const affiliateFee = rev * (row.affiliateRate / 100);
                const other = row.other || 0;
                
                totalRevenue += rev;
                totalFees += fees;
                totalCogsAndOther += cogs + affiliateFee + other;
                totalProfit += (rev - (cogs + affiliateFee + other + fees));
            }});
            
            const margin = totalRevenue > 0 ? (totalProfit / totalRevenue) * 100 : 0;
            
            document.getElementById('tt-dash-revenue').innerText = Math.round(totalRevenue).toLocaleString('vi-VN') + " VNĐ";
            document.getElementById('tt-dash-fees').innerText = Math.round(totalFees).toLocaleString('vi-VN') + " VNĐ";
            document.getElementById('tt-dash-cogs').innerText = Math.round(totalCogsAndOther).toLocaleString('vi-VN') + " VNĐ";
            
            const profitEl = document.getElementById('tt-dash-profit');
            profitEl.innerText = Math.round(totalProfit).toLocaleString('vi-VN') + " VNĐ";
            profitEl.style.color = totalProfit >= 0 ? '#22c55e' : '#ef4444';
            
            document.getElementById('tt-dash-margin').innerText = margin.toFixed(1) + "%";
        }}

        // --- SHOPEE DB ---
        function renderShopeeDB() {{
            const tbody = document.getElementById('shopee-db-tbody');
            if (!tbody) return;
            tbody.innerHTML = '';
            
            shopeeProducts.forEach((p, idx) => {{
                const tr = document.createElement('tr');
                tr.innerHTML = `
                    <td style="padding: 6px; font-weight:600;">${{p.sku}}</td>
                    <td style="padding: 6px;">${{p.name}}</td>
                    <td style="padding: 6px; text-align:right;">${{(p.comm_regular * 100).toFixed(1)}}%</td>
                    <td style="padding: 6px; text-align:right;">${{(p.comm_mall * 100).toFixed(1)}}%</td>
                    <td style="padding: 6px; text-align:right;">${{p.cogs !== null ? Math.round(p.cogs).toLocaleString() : '-'}}đ</td>
                    <td style="padding: 6px; text-align:right;">${{Math.round(p.packing).toLocaleString()}}đ</td>
                    <td style="padding: 6px; text-align:center;">
                        <button onclick="deleteShopeeDBProduct(${{"'" + p.sku + "'"}})" style="background:none; border:none; color:#ef4444; cursor:pointer; font-weight:bold;">Xóa</button>
                    </td>
                `;
                tbody.appendChild(tr);
            }});
        }}

        function searchShopeeDB() {{
            const query = document.getElementById('shopee-db-search').value.toLowerCase();
            const tbody = document.getElementById('shopee-db-tbody');
            tbody.querySelectorAll('tr').forEach(tr => {{
                const text = tr.innerText.toLowerCase();
                tr.style.display = text.includes(query) ? '' : 'none';
            }});
        }}

        function openAddShopeeProductModal() {{
            document.getElementById('shopee-add-modal').classList.add('show');
        }}

        function closeShopeeAddModal() {{
            document.getElementById('shopee-add-modal').classList.remove('show');
        }}

        function saveCustomShopeeProduct() {{
            const sku = document.getElementById('add-sp-sku').value.trim();
            const name = document.getElementById('add-sp-name').value.trim();
            const price = parseFloat(document.getElementById('add-sp-price').value) || 0;
            const reg = (parseFloat(document.getElementById('add-sp-comm-regular').value) || 10) / 100;
            const mall = (parseFloat(document.getElementById('add-sp-comm-mall').value) || 12.5) / 100;
            const cogs = parseFloat(document.getElementById('add-sp-cogs').value) || 0;
            const packing = parseFloat(document.getElementById('add-sp-packing').value) || 0;
            
            if (!sku || !name) {{
                alert("Vui lòng điền đầy đủ mã SKU và tên sản phẩm!");
                return;
            }}
            
            if (shopeeProducts.some(p => p.sku === sku)) {{
                alert("Mã SKU này đã tồn tại trong danh mục!");
                return;
            }}
            
            shopeeProducts.push({{
                sku: sku,
                name: name,
                brand: "Tùy chỉnh",
                category: "Custom",
                comm_regular: reg,
                comm_mall: mall,
                ref_price: price,
                cogs: cogs,
                packing: packing,
                default_cost: cogs + packing,
                policy: "Sản phẩm tự thêm"
            }});
            
            localStorage.setItem('shopee_product_db', JSON.stringify(shopeeProducts));
            closeShopeeAddModal();
            renderShopeeDB();
            renderShopeeTable();
            alert("Đã thêm sản phẩm thành công!");
        }}

        function deleteShopeeDBProduct(sku) {{
            if (!confirm("Bạn có chắc chắn muốn xóa sản phẩm này ra khỏi cơ sở dữ liệu?")) return;
            shopeeProducts = shopeeProducts.filter(p => p.sku !== sku);
            localStorage.setItem('shopee_product_db', JSON.stringify(shopeeProducts));
            renderShopeeDB();
            
            shopeeCalcRows = shopeeCalcRows.filter(r => r.sku !== sku);
            if (shopeeCalcRows.length === 0 && shopeeProducts.length > 0) {{
                shopeeCalcRows.push({{
                    sku: shopeeProducts[0].sku,
                    shopType: 'regular',
                    price: shopeeProducts[0].ref_price || 200000,
                    voucher: 0,
                    affiliateRate: 0,
                    cogs: null,
                    other: 0
                }});
            }}
            localStorage.setItem('shopee_calc_rows', JSON.stringify(shopeeCalcRows));
            renderShopeeTable();
        }}

        function resetShopeeDB() {{
            if (!confirm("Khôi phục danh mục sản phẩm Shopee về mặc định của file Excel?")) return;
            shopeeProducts = [...DEFAULT_SHOPEE_PRODUCTS];
            localStorage.setItem('shopee_product_db', JSON.stringify(shopeeProducts));
            renderShopeeDB();
            
            shopeeCalcRows = [{{
                sku: shopeeProducts[0].sku,
                shopType: 'regular',
                price: shopeeProducts[0].ref_price || 200000,
                voucher: 0,
                affiliateRate: 0,
                cogs: null,
                other: 0
            }}];
            localStorage.setItem('shopee_calc_rows', JSON.stringify(shopeeCalcRows));
            renderShopeeTable();
        }}

        function exportShopeeDB() {{
            const dataStr = "data:text/json;charset=utf-8," + encodeURIComponent(JSON.stringify(shopeeProducts, null, 2));
            const dlAnchorElem = document.createElement('a');
            dlAnchorElem.setAttribute("href", dataStr);
            dlAnchorElem.setAttribute("download", "shopee_products_catalog.json");
            dlAnchorElem.click();
        }}

        function importShopeeDB() {{
            const input = document.createElement('input');
            input.type = 'file';
            input.accept = '.json';
            input.onchange = e => {{
                const file = e.target.files[0];
                const reader = new FileReader();
                reader.readAsText(file,'UTF-8');
                reader.onload = readerEvent => {{
                    const content = readerEvent.target.result;
                    try {{
                        const parsed = JSON.parse(content);
                        if (Array.isArray(parsed)) {{
                            shopeeProducts = parsed;
                            localStorage.setItem('shopee_product_db', JSON.stringify(shopeeProducts));
                            renderShopeeDB();
                            renderShopeeTable();
                            alert("Nhập danh mục JSON thành công!");
                        }} else {{
                            alert("Định dạng file JSON không hợp lệ (Phải là một mảng)!");
                        }}
                    }} catch(err) {{
                        alert("Lỗi khi đọc file JSON!");
                    }}
                }}
            }}
            input.click();
        }}

        // --- TIKTOK DB ---
        function renderTikTokDB() {{
            const tbody = document.getElementById('tiktok-db-tbody');
            if (!tbody) return;
            tbody.innerHTML = '';
            
            tiktokProducts.forEach((p, idx) => {{
                const tr = document.createElement('tr');
                tr.innerHTML = `
                    <td style="padding: 6px; font-weight:600;">${{p.sku}}</td>
                    <td style="padding: 6px;">${{p.name}}</td>
                    <td style="padding: 6px; text-align:right;">${{(p.comm_regular * 100).toFixed(1)}}%</td>
                    <td style="padding: 6px; text-align:right;">${{(p.comm_mall * 100).toFixed(1)}}%</td>
                    <td style="padding: 6px; text-align:right;">${{Math.round(p.cogs).toLocaleString()}}đ</td>
                    <td style="padding: 6px; text-align:right;">${{Math.round(p.packing).toLocaleString()}}đ</td>
                    <td style="padding: 6px; text-align:center;">
                        <button onclick="deleteTikTokDBProduct(${{"'" + p.sku + "'"}})" style="background:none; border:none; color:#ef4444; cursor:pointer; font-weight:bold;">Xóa</button>
                    </td>
                `;
                tbody.appendChild(tr);
            }});
        }}

        function searchTikTokDB() {{
            const query = document.getElementById('tiktok-db-search').value.toLowerCase();
            const tbody = document.getElementById('tiktok-db-tbody');
            tbody.querySelectorAll('tr').forEach(tr => {{
                const text = tr.innerText.toLowerCase();
                tr.style.display = text.includes(query) ? '' : 'none';
            }});
        }}

        function openAddTikTokProductModal() {{
            document.getElementById('tiktok-add-modal').classList.add('show');
        }}

        function closeTikTokAddModal() {{
            document.getElementById('tiktok-add-modal').classList.remove('show');
        }}

        function saveCustomTikTokProduct() {{
            const sku = document.getElementById('add-tt-sku').value.trim();
            const name = document.getElementById('add-tt-name').value.trim();
            const reg = (parseFloat(document.getElementById('add-tt-comm-regular').value) || 10) / 100;
            const mall = (parseFloat(document.getElementById('add-tt-comm-mall').value) || 12.5) / 100;
            const cogs = parseFloat(document.getElementById('add-tt-cogs').value) || 0;
            const packing = parseFloat(document.getElementById('add-tt-packing').value) || 0;
            
            if (!sku || !name) {{
                alert("Vui lòng điền đầy đủ mã SKU và tên sản phẩm!");
                return;
            }}
            
            if (tiktokProducts.some(p => p.sku === sku)) {{
                alert("Mã SKU này đã tồn tại trong danh mục!");
                return;
            }}
            
            tiktokProducts.push({{
                sku: sku,
                name: name,
                brand: "Tùy chỉnh",
                category: "Custom",
                comm_regular: reg,
                comm_mall: mall,
                cogs: cogs,
                packing: packing,
                gift: 0
            }});
            
            localStorage.setItem('tiktok_product_db', JSON.stringify(tiktokProducts));
            closeTikTokAddModal();
            renderTikTokDB();
            renderTikTokTable();
            alert("Đã thêm sản phẩm thành công!");
        }}

        function deleteTikTokDBProduct(sku) {{
            if (!confirm("Bạn có chắc chắn muốn xóa sản phẩm này ra khỏi cơ sở dữ liệu?")) return;
            tiktokProducts = tiktokProducts.filter(p => p.sku !== sku);
            localStorage.setItem('tiktok_product_db', JSON.stringify(tiktokProducts));
            renderTikTokDB();
            
            tiktokCalcRows = tiktokCalcRows.filter(r => r.sku !== sku);
            if (tiktokCalcRows.length === 0 && tiktokProducts.length > 0) {{
                tiktokCalcRows.push({{
                    sku: tiktokProducts[0].sku,
                    shopType: 'regular',
                    price: 200000,
                    voucher: 0,
                    affiliateRate: 0,
                    cogs: null,
                    other: 0
                }});
            }}
            localStorage.setItem('tiktok_calc_rows', JSON.stringify(tiktokCalcRows));
            renderTikTokTable();
        }}

        function resetTikTokDB() {{
            if (!confirm("Khôi phục danh mục sản phẩm TikTok Shop về mặc định?")) return;
            tiktokProducts = [...DEFAULT_TIKTOK_PRODUCTS];
            localStorage.setItem('tiktok_product_db', JSON.stringify(tiktokProducts));
            renderTikTokDB();
            
            tiktokCalcRows = [{{
                sku: tiktokProducts[0].sku,
                shopType: 'regular',
                price: 200000,
                voucher: 0,
                affiliateRate: 0,
                cogs: null,
                other: 0
            }}];
            localStorage.setItem('tiktok_calc_rows', JSON.stringify(tiktokCalcRows));
            renderTikTokTable();
        }}

        function exportTikTokDB() {{
            const dataStr = "data:text/json;charset=utf-8," + encodeURIComponent(JSON.stringify(tiktokProducts, null, 2));
            const dlAnchorElem = document.createElement('a');
            dlAnchorElem.setAttribute("href", dataStr);
            dlAnchorElem.setAttribute("download", "tiktok_products_catalog.json");
            dlAnchorElem.click();
        }}

        function importTikTokDB() {{
            const input = document.createElement('input');
            input.type = 'file';
            input.accept = '.json';
            input.onchange = e => {{
                const file = e.target.files[0];
                const reader = new FileReader();
                reader.readAsText(file,'UTF-8');
                reader.onload = readerEvent => {{
                    const content = readerEvent.target.result;
                    try {{
                        const parsed = JSON.parse(content);
                        if (Array.isArray(parsed)) {{
                            tiktokProducts = parsed;
                            localStorage.setItem('tiktok_product_db', JSON.stringify(tiktokProducts));
                            renderTikTokDB();
                            renderTikTokTable();
                            alert("Nhập danh mục JSON thành công!");
                        }} else {{
                            alert("Định dạng file JSON không hợp lệ (Phải là một mảng)!");
                        }}
                    }} catch(err) {{
                        alert("Lỗi khi đọc file JSON!");
                    }}
                }}
            }}
            input.click();
        }}

        // --- Init rendering ---
        renderShopeeTable();
        renderShopeeDB();
        renderTikTokTable();
        renderTikTokDB();
    </script>
</body>
</html>
"""

# Write output file
with open(output_html, "w", encoding="utf-8") as f:
    f.write(template_head + body_content + template_foot)

print("Unified handbook compiled successfully with advanced multi-row calculators and product database manager!")
